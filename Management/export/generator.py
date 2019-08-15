from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

class ExcelColumn:
    def __init__(self, title, style=None, showSum=False, width=None, columns=None):
        self.title = title
        self.style = style
        self.showSum = showSum
        self.width = width
        self.columns = columns

class ExcelGenerator:

    def __init__(self):

        self.thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin'))

    def _create_workbook(self):
        workbook = Workbook()
        # Get active worksheet/tab
        worksheet = workbook.active
        # set RTL
        worksheet.sheet_view.rightToLeft = True
        
        self.workbook = workbook
        self.worksheet = worksheet

    def _set_col_width(self, col_num, width):
        col_letter_ascii = ord('A') + col_num
        col_letter = chr(col_letter_ascii)

        self.worksheet.column_dimensions[col_letter].width = width

    def _size_columns(self, columns):

        col_num = 0

        for col in columns:
            sub_columns = col.columns

            if sub_columns == None:
                if col.width != None:
                    self._set_col_width(col_num, col.width)

                col_num += 1
            else:
                for sub_col in sub_columns:
                    if sub_col.width != None:
                        self._set_col_width(col_num, sub_col.width)
                    
                    col_num += 1

    def _create_title(self, title):
        title_cell = self.worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.style = 'Headline 1'
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        col_span = len(self._columns)

        self.worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=col_span)

    def _create_header_cell(self, row, column, value):
        cell = self.worksheet.cell(row=row, column=column)
        cell.value = value

        cell.style = 'Accent3'
        cell.border = self.thin_border

    def _create_table_headers(self, columns):
        row_num = 2
        col_num = 1
        has_sub_columns = False

        actual_columns = []

        for column in columns:
            self._create_header_cell(row_num, col_num, column.title)

            if column.columns == None:
                actual_columns.append(column)
                col_num += 1
            else:
                has_sub_columns = True

                for sub_col_num, sub_col in enumerate(column.columns):
                    self._create_header_cell(
                        row_num + 1, 
                        col_num + sub_col_num, 
                        sub_col.title)

                    actual_columns.append(sub_col)

                end_column = col_num + len(column.columns) - 1

                # merge parent column
                self.worksheet.merge_cells(
                    start_row=row_num,
                    start_column=col_num,
                    end_row=row_num,
                    end_column=end_column)

                col_num = end_column + 1

        # set row_num
        self.row_num = 3 if has_sub_columns else 2
        self._columns = actual_columns

    def _create_table_rows(self, data_rows):
        sum_row = [0 if col.showSum else '' for col in self._columns]
        # override first cell
        sum_row[0] = 'סה"כ'

        for row in data_rows:
            self.row_num += 1

            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = self.worksheet.cell(row=self.row_num, column=col_num)
                cell.value = cell_value if cell_value != None else ''

                col = self._columns[col_num - 1]

                if col.style == 'currency':
                    cell.style = 'Currency'
                    cell.number_format = '#,##0 ₪'
                elif col.style == 'percent':
                    cell.style = 'Percent'
                    # scale cell_value
                    cell.value = cell_value / 100 if cell_value else ''

                cell.border = self.thin_border

                if col.showSum and cell_value != None:
                    sum_row[col_num - 1] += cell_value

        self.sum_row = sum_row

    def _create_summary_row(self):
        row_num = self.row_num + 1

        for col_num, cell_value in enumerate(self.sum_row, 1):
            cell = self.worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value or ''

            col = self._columns[col_num - 1]

            if col.style == 'currency':
                cell.style = 'Currency'
                cell.number_format = '#,##0 ₪'

            cell.font = Font(bold=True)
            cell.border = self.thin_border


    def generate(self, title, columns, data_rows):
        self.title = title
        self.columns = columns
        self.data_rows = data_rows

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={title}.xlsx'.format(
            title=title,
        )

        self._create_workbook()

        self._size_columns(columns)

        self._create_table_headers(columns)

        self._create_title(title)

        self._create_table_rows(data_rows)

        self._create_summary_row()

        self.workbook.save(response)

        return response
