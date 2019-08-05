import itertools, time, io
from datetime import datetime, date

import django.core.paginator
from django.db.models import Count, Sum, Q
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseBadRequest, HttpResponseServerError, FileResponse
from django.forms.models import inlineformset_factory
from django.forms.formsets import formset_factory
from django.core import serializers
from django.urls import reverse

from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.contenttypes.models import ContentType

from Indices.models import Tax

import Management.common as common
from .forms import *
from .models import *
from .pdf.writers import MonthDemandWriter, MultipleDemandWriter, EmployeeListWriter, EmployeeSalariesWriter, ProjectListWriter, EmployeesLoans
from .pdf.writers import PricelistWriter, BuildingClientsWriter, EmployeeSalariesBookKeepingWriter, SalariesBankWriter, DemandFollowupWriter
from .pdf.writers import EmployeeSalesWriter, DemandPayBalanceWriter

from .enrichers.demand import set_demand_diff_fields, set_demand_is_fixed, set_demand_last_status
from .enrichers.demand import set_demand_open_reminders, set_demand_invoice_payment_fields
from .enrichers.demand import set_demand_sale_fields

from .enrichers.salary import enrich_employee_salaries, enrich_nh_employee_salaries
from .enrichers.salary import set_salary_base_fields, set_employee_sales, set_salary_status_date
from .enrichers.salary import set_loan_fields

from .mail import mail
from pprint import pprint


def build_and_return_pdf(writer):
    # Create a file-like buffer to receive PDF data.
    buffer = io.BytesIO()

    writer.build(buffer)

    buffer.seek(io.SEEK_SET)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    return FileResponse(buffer, as_attachment=True, filename='hello.pdf')  
    
@login_required
def index(request):
    context = {
        'locateHouseForm':LocateHouseForm(), 
        'locateDemandForm': LocateDemandForm(),
        'employeeSalesForm': ProjectSeasonForm(), 
        'employeeSalarySeasonForm': EmployeeSeasonForm(),
        'nhbranches':NHBranch.objects.all()
        }
    return render(request, 'Management/index.html', context)
  
@login_required  
def locate_house(request):
    from django.core.exceptions import ObjectDoesNotExist

    if request.method == 'GET':
        form = LocateHouseForm(request.GET)
        if form.is_valid():
            project = form.cleaned_data['project']
            try:
                building = project.buildings.get(num=form.cleaned_data['building_num'])
                house = building.houses.get(num=form.cleaned_data['house_num'])
                return HttpResponseRedirect('/buildings/%s/house/%s/type1' % (building.id, house.id))
            except ObjectDoesNotExist:
                error = u'לא נמצאה דירה מס %s בבניין מס %s בפרוייקט %s' % (form.cleaned_data['house_num'],
                                                                           form.cleaned_data['building_num'],
                                                                           project)
                return render(request, 'Management/error.html', {'error': error})

    return HttpResponseRedirect('/')
    
@login_required  
def locate_demand(request):
    if request.method == 'GET':
        form = LocateDemandForm(request.GET)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            try:
                demand = Demand.objects.get(**cleaned_data)
                
                if 'find' in request.GET:
                    return HttpResponseRedirect(demand.get_absolute_url())
                elif 'pdf' in request.GET:
                    return report_project_month(request, demand = demand)
                
            except Demand.DoesNotExist:
                error = u'לא נמצאה דרישה מתאימה'
                return render(request, 'Management/error.html', {'error': error})

    return HttpResponseRedirect('/')
    
@login_required
def employee_loans(request, object_id):

    employee_base = EmployeeBase.objects.get(pk=object_id)

    # set to Employee or NHEmployee
    employee = employee_base.derived

    set_loan_fields([employee])

    if request.GET.get('t') == 'pdf' :
        writer = EmployeesLoans(employee)
        return build_and_return_pdf(writer)
    else:
        return render(request, 'Management/employee_loans.html', {'employee':employee})

@login_required
def limited_update_object(request, permission=None, *args, **kwargs):
    if 'model' in kwargs:
        model = kwargs['model']
    elif 'form_class' in kwargs:
        model = kwargs['form_class']._meta.model

    model_name = model.__name__.lower()

    if request.method == "POST" :
        if model_name == 'projectcommission':
            allow_history = [ 'add_amount', 'add_type', 'registration_amount', 'max' ]
            form_old      = ProjectCommission()
            form_old_dict = form_old.__dict__.copy()
            form          = ProjectCommissionForm(request.POST, instance=form_old)

            if form.is_valid():
                form_dict = form.instance.__dict__

                for item in form_old._meta.fields :
                    name = item.name

                    if name in form_dict and name in allow_history :
                        value     = form_dict.get(name)
                        value_len = value != None and len(str(value)) or 0

                        if value_len > 0 and ( form_old_dict.get(name) != value ):
                            TransactionUpdateHistory( transaction_type='1', transaction_id=kwargs.get('object_id'), field_name=name, field_value=(value and value or ''), timestamp=datetime.now() ).save()

    if not permission:
        permission = 'Management.change_' + model_name
    if request.user.has_perm(permission):
        return UpdateView.as_view(*args, **kwargs)
    else:
        return HttpResponse('No permission. contact Elad.')

class HouseDetailView(LoginRequiredMixin, DetailView):
    model = House
    context_object_name = 'house'
    template_name = 'Management/house_details.html'
    
class SignupDetailView(LoginRequiredMixin, DetailView):
    model = Signup
    context_object_name = 'signup'
    template_name = 'Management/signup_details.html'

### Loan Views ###

class LoanListView(PermissionRequiredMixin, ListView):
    model = Loan
    permission_required = 'Management.list_loan'

    def get_queryset(self):
        return Loan.objects.all()

class LoanCreate(PermissionRequiredMixin, CreateView):
    model = Loan
    form_class = LoanForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.add_loan'

class LoanUpdate(PermissionRequiredMixin, UpdateView):
    model = Loan
    form_class = LoanForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_loan'

class LoanDelete(PermissionRequiredMixin, DeleteView):
    model = Loan
    success_url = '/loans'
    template_name = 'Management/object_confirm_delete.html'
    permission_required = 'Management.delete_loan'

### Lawer Views ###

class LawyerListView(LoginRequiredMixin, ListView):
    model = Lawyer

    def get_queryset(self):
        return Lawyer.objects.all()

class LawyerCreate(PermissionRequiredMixin, CreateView):
    model = Lawyer
    fields = ('first_name','last_name','cell_phone','mail','address','role','phone')
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.add_lawyer'

class LawyerUpdate(PermissionRequiredMixin, UpdateView):
    model = Lawyer
    fields = ('first_name','last_name','cell_phone','mail','address','role','phone')
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_lawyer'

class LawyerDelete(PermissionRequiredMixin, DeleteView):
    model = Lawyer
    success_url = '/lawyers'
    template_name = 'Management/object_confirm_delete.html'
    permission_required = 'Management.delete_lawyer'

### NHCommission Views ###

class NHCommissionCreate(PermissionRequiredMixin, CreateView):
    model = NHCommission
    form_class = NHCommissionForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.add_nhcommission'

class NHCommissionUpdate(PermissionRequiredMixin, UpdateView):
    model = NHCommission
    form_class = NHCommissionForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_nhcommission'

class NHCommissionDelete(PermissionRequiredMixin, DeleteView):
    model = NHCommission
    success_url = '/nhcommissions'
    template_name = 'Management/object_confirm_delete.html'
    permission_required = 'Management.delete_nhcommission'

### Demand Views ###

class DemandRemarksUpdate(PermissionRequiredMixin, UpdateView):
    model = Demand
    form_class = DemandRemarksForm
    template_name = 'Management/object_edit.html'
    success_url = 'remarks'
    permission_required = 'Management.demand_remarks'
    
class DemandSaleCountUpdate(PermissionRequiredMixin, UpdateView):
    model = Demand
    form_class = DemandSaleCountForm
    template_name = 'Management/object_edit.html'
    success_url = 'salecount'
    permission_required = 'Management.demand_sale_count'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        
        context['title'] = u"עדכון מס' מכירות צפוי"
        return context

### Account ###

class AccountUpdate(PermissionRequiredMixin, UpdateView):
    model = Account
    form_class = AccountForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_account'

### EPCommission ###

class EPCommissionUpdate(PermissionRequiredMixin, UpdateView):
    model = EPCommission
    fields = ('start_date','end_date','max')
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_epcommission'

### NHEmployee ###

class NHEmployeeUpdate(PermissionRequiredMixin, UpdateView):
    model = NHEmployee
    form_class = NHEmployeeForm
    permission_required = 'Management.change_nhemployee'

class NHEmployeeEndUpdate(PermissionRequiredMixin, UpdateView):
    model = NHEmployee
    form_class = EmployeeEndForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_nhemployee'

def signup_list(request, project_id):
    month = date.today()
    y = int(request.GET.get('year', month.year))
    m = int(request.GET.get('month', month.month))
    month = datetime(y,m,1)
    form = MonthForm(initial={'year':y,'month':m})
    p = Project.objects.get(pk = project_id)
    signups = p.signups(y, m)
    return render(request, 'Management/signup_list.html', 
                          { 'project':p, 'signups':signups, 'month':month, 'filterForm':form },
                          )

@permission_required('Management.add_signup')
def signup_edit(request, id=None, house_id=None, project_id=None): 
    if request.method == 'POST':
        form = SignupForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('add')
    else:
        if id:
            signup = Signup.objects.get(pk=id)
            h = signup.house
            form = SignupForm(instance=signup)
        elif house_id:
            h = House.objects.get(pk= house_id)  
            signup = h.get_signup()
            form = signup and SignupForm(instance = signup) or SignupForm()
        if house_id or id:
            form.fields['house'].initial = h.id
            form.fields['building'].initial = h.building.id
            form.fields['project'].initial = h.building.project.id
            form.fields['house'].queryset = h.building.houses.all()
            form.fields['building'].queryset = h.building.project.non_deleted_buildings()
        elif project_id:
            form = SignupForm(initial = {'project':project_id})
            form.fields['building'].queryset = Building.objects.filter(project__id = project_id, is_deleted=False)
        else:
            form = SignupForm()
            
    return render(request, 'Management/signup_edit.html', 
                              { 'form':form },
                              )

@permission_required('Management.change_signup')
def signup_cancel(request, id):
    s = Signup.objects.get(pk=id)
    cancel = s.cancel or SignupCancel()
    if request.method=='POST':
        form = SignupCancelForm(request.POST)
        if form.is_valid():
            s.cancel = form.save()
            s.save()
    else:
        form = SignupCancelForm(instance=cancel)
    return render(request, 'Management/object_edit.html', 
                              { 'form':form },
                              )

@permission_required('Management.add_account')
def employee_account(request, id, model):
    employee = model.objects.get(pk=id)
    try:
        acc = employee.account
    except Account.DoesNotExist:
        acc = Account()
    if request.method == 'POST':
        form = AccountForm(data=request.POST, instance=acc) 
        if form.is_valid(): 
            employee.account = form.save()
            employee.save()
    else:
        form = AccountForm(instance=acc)
        
    return render(request, 'Management/object_edit.html', 
                              { 'form':form },
                              )

@permission_required('Management.list_demand')
def demand_function(request,id , function):
    d = Demand.objects.get(pk=id)
    function(d)
    return HttpResponse('ok')

@permission_required('Management.list_demand')
def demand_return_to_calc(request, id):
    demand = Demand.objects.get(pk = id)
    demand.close()
    url = reverse('demand-old') + '?year=%s&month=%s' % (demand.year,demand.month)
    return HttpResponseRedirect(url)

@permission_required('Management.list_demand')
def demand_calc(request, id):
    # end the revision created by the middleware - i want manual control of the revision
    #reversion.revision.end()
        
    d = Demand.objects.get(pk=id)
    commissions = d.project.commissions.get()

    if commissions.commission_by_signups or commissions.c_zilber:
        if commissions.commission_by_signups:
            demands = list(Demand.objects.filter(project = d.project))
        elif commissions.c_zilber:
            demand = d
            demands = []
            while demand.zilber_cycle_index() > 1:
                demands.insert(0, demand)
                demand = demand.get_previous_demand()
            demands.insert(0, demand)
        
        # enrich demands
        set_demand_diff_fields(demands)
        set_demand_last_status(demands)

        # exclude all demands that were already sent! to include them you must manually change their status!!!!!
        demands = [demand for demand in demands if demand.last_status.type_id not in (DemandStatusType.Sent, DemandStatusType.Finished)]
            
        # delete all commissions and sale commission details before re-calculating
        for demand in demands:
            # enrich demand
            year, month = demand.year, demand.month
            set_demand_sale_fields([demand], year, month, year, month)

            # delete demand statuses
            demand.statuses.delete()

        for d2 in demands:
            d2.calc_sales_commission()
            demand = Demand.objects.get(pk=d2.id)
            if demand.get_next_demand() != None:
                demand.finish()
                time.sleep(1)
    else:
        # enrich demand
        set_demand_sale_fields([d], d.year, d.month, d.year, d.month)
        set_demand_diff_fields([d])

        # re-calculate commission for demand
        d.calc_sales_commission()
        
    url = reverse('demand-old') + '?year=%s&month=%s' % (d.year,d.month)
    return HttpResponseRedirect(url)

@permission_required('Management.list_demand')
def demand_old_list(request):
    year, month = None, None
    ds, unhandled_projects = [], []
    total_sales_count,total_sales_amount, total_amount = 0,0,0
        
    if len(request.GET):
        form = MonthForm(request.GET)
        if form.is_valid():
            year, month = form.cleaned_data['year'], form.cleaned_data['month']
    else:
        current = common.current_month()
        year, month = current.year, current.month
        form = MonthForm(initial={'year':year,'month':month})
        
    if year and month:
        ds = Demand.objects \
            .filter(year = year, month = month) \
            .annotate(Count('statuses')) \
            .prefetch_related('reminders__statuses') \
            .select_related('project')

        unhandled_projects.extend(Project.objects.active())

        set_demand_sale_fields(ds, year, month, year, month)
        set_demand_diff_fields(ds)
        set_demand_is_fixed(ds)
        set_demand_last_status(ds)
        set_demand_open_reminders(ds)

        for d in ds:
            total_sales_count += d.sales_count
            total_sales_amount += d.sales_amount
            total_amount += d.total_amount

            if d.last_status and d.last_status.type_id in [DemandStatusType.Sent, DemandStatusType.Finished]:
                try:
                    unhandled_projects.remove(d.project)
                except ValueError:
                    pass
        
    context =  { 
        'demands':ds, 
        'month':date(year, month, 1),
        'filterForm':form,
        'total_sales_count':total_sales_count,
        'total_sales_amount':total_sales_amount,
        'total_amount':total_amount,
        'unhandled_projects':unhandled_projects
    }

    return render(request, 'Management/demand_old_list.html', context)

def nhemployee_salary_send(request, nhbranch_id, year, month):
    pass
    
def nhemployee_salary_pdf(request, nhbranch_id, year, month):
    nhb = NHBranch.objects.get(pk = nhbranch_id)

    salaries = NHEmployeeSalary.objects \
        .select_related('nhemployee__employment_terms__hire_type') \
        .nondeleted() \
        .filter(nhbranch = nhb, year = year, month= month) \
        .order_by('id')

    employee_by_id = {s.nhemployee.id:s.nhemployee for s in salaries}

    set_salary_base_fields(
        salaries, 
        employee_by_id, 
        year, month, year, month)

    set_salary_status_date(salaries)

    approved_salaries = [salary for salary in salaries if salary.approved_date]

    nhsales = NHSale.objects.filter(nhmonth__year__exact = year, nhmonth__month__exact = month, nhmonth__nhbranch = nhb)
    title = u'שכר עבודה לסניף %s לחודש %s\%s' % (nhb, year, month)

    writer = EmployeeSalariesBookKeepingWriter(approved_salaries, title, nhsales)
    
    return build_and_return_pdf(writer)

@permission_required('Management.change_salaryexpenses')
def employee_salary_expenses(request, salary_id):
    salary = EmployeeSalaryBase.objects.get(pk=salary_id)
    
    # set to EmployeeSalary or NHEmployeeSalary
    salary = salary.derived

    employee = salary.get_employee()
    terms = employee.employment_terms

    year, month = salary.year, salary.month

    set_salary_base_fields(
        [salary],
        {employee.id:employee},
        year, month, year, month)

    expenses = salary.expenses or SalaryExpenses(employee = employee, year = year, month = month)

    if request.method=='POST':
        form = SalaryExpensesForm(request.POST, instance= expenses)
        if form.is_valid():
            form.save()
    else:
        vacation = terms.salary_base and (terms.salary_base / 24) or (2500/12)
        form = SalaryExpensesForm(instance= expenses, initial={'vacation':vacation})

    context = {'form': form, 'neto': salary.neto or 0}
    
    return render(request, 'Management/salaryexpenses_edit.html', context)

@permission_required('Management.change_employeesalary')
def employee_salary_approve(request, id):
    es = EmployeeSalaryBase.objects.get(pk=id)
    es.approve()

    if hasattr(es,'employeesalary'):
        url = reverse('salary-list') + '?year=%s&month=%s' % (es.year, es.month)
    elif hasattr(es,'nhemployeesalary'):
        url = reverse('nh-salary-list') + '?year=%s&month=%s' % (es.year, es.month)

    return HttpResponseRedirect(url)

@permission_required('Management.change_employeesalary')
def salary_expenses_approve(request, id):
    se = SalaryExpenses.objects.get(pk=id)
    se.approve()
    se.save()
    url = reverse('salary-expenses') + '?year=%s&month=%s' % (se.year, se.month)
    return HttpResponseRedirect(url)
    
class SalaryExpensesUpdate(PermissionRequiredMixin, UpdateView):
    model = SalaryExpenses
    exclude = ('approved_date',)
    template_name = 'Management/salaryexpenses_edit.html'
    permission_required = 'Management.change_salaryexpenses'

@permission_required('Management.list_employeesalary')
def employee_salary_list(request):
    current = common.current_month()
    year = int(request.GET.get('year', current.year))
    month = int(request.GET.get('month', current.month))
    salaries = []
    today = date.today()
    if date(year, month, 1) <= today:
        employees = Employee.objects \
            .select_related('employment_terms__hire_type','rank') \
            .prefetch_related('projects') \
            .filter(employment_terms__isnull=False)

        for e in employees:
            # do not include employees who did not start working by the month selected
            if year < e.work_start.year or (year == e.work_start.year and month < e.work_start.month):
                continue
            # do not include employees who finished working by the month selected
            if e.work_end and (year > e.work_end.year or (year == e.work_end.year and month > e.work_end.month)):
                continue
            es, new = EmployeeSalary.objects.get_or_create(employee = e, month = month, year = year)
            if new:
                es.calculate()
                es.save()
            else:
                if es.is_deleted:
                    continue

            # set employee field with already-loaded fields
            es.employee = e

            salaries.append(es)
        
        employee_by_id = {employee.id:employee for employee in employees}

        enrich_employee_salaries(salaries, employee_by_id, year, month, year, month)

        set_loan_fields(employees)

    context = {
        'salaries':salaries, 
        'month': date(int(year), int(month), 1),
        'filterForm':MonthForm(initial={'year':year,'month':month})
        }
    
    return render(request, 'Management/employee_salaries.html', context)

class SalaryExpensesListView(PermissionRequiredMixin, ListView):
    model = EmployeeSalary
    template_name = 'Management/salaries_expenses.html'
    context_object_name = 'salaries'
    permission_required = 'Management.list_salaryexpenses'

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)

        current = common.current_month()
        self.year = int(self.request.GET.get('year', current.year))
        self.month = int(self.request.GET.get('month', current.month))

    def get_queryset(self):
        salaries = EmployeeSalary.objects.nondeleted() \
            .select_related('employee__employment_terms__hire_type', 'employee__rank') \
            .prefetch_related('employee__projects') \
            .filter(year = self.year, month = self.month)

        employee_by_id = {s.employee.id:s.employee for s in salaries}

        enrich_employee_salaries(salaries, employee_by_id, self.year, self.month, self.year, self.month)

        return salaries

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        
        context['month'] = date(int(self.year), int(self.month), 1)
        context['filterForm'] = MonthForm(initial={'year':self.year,'month':self.month})

        return context

class NHSalaryExpensesListView(PermissionRequiredMixin, ListView):
    model = EmployeeSalary
    template_name = 'Management/salaries_expenses.html'
    context_object_name = 'salaries'
    permission_required = 'Management.nh_salaries_expenses'

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)

        current = common.current_month()
        self.year = int(self.request.GET.get('year', current.year))
        self.month = int(self.request.GET.get('month', current.month))

    def get_queryset(self):
        salaries = NHEmployeeSalary.objects.nondeleted() \
            .select_related('nhemployee__employment_terms__hire_type') \
            .filter(year = self.year, month = self.month)

        employee_by_id = {s.nhemployee.id:s.nhemployee for s in salaries}

        enrich_nh_employee_salaries(salaries, employee_by_id, self.year, self.month, self.year, self.month)

        return salaries

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        
        context['month'] = date(int(self.year), int(self.month), 1)
        context['filterForm'] = MonthForm(initial={'year':self.year,'month':self.month})

        return context

@permission_required('Management.list_nhemployeesalary')
def nhemployee_salary_list(request):
    current = common.current_month()
    year = int(request.GET.get('year', current.year))
    month = int(request.GET.get('month', current.month))
    
    salaries = []

    # key is NHBranch, value is a list of NHEmployeeSalary objects
    branch_list = {}

    for nhbe in NHBranchEmployee.objects.month(year, month):
        es, new = NHEmployeeSalary.objects.get_or_create(nhemployee = nhbe.nhemployee, nhbranch = nhbe.nhbranch,
                                                         month = month, year = year)
        if not new and es.is_deleted:
            continue
        #if (new or not es.commissions or not es.base or not es.admin_commission) and es.approved_date == None: 
        if new:
            es.calculate()
            es.save()
        
        salaries.append(es)

        branch_sales = branch_list.setdefault(nhbe.nhbranch, [])
        branch_sales.append(es)

    employee_by_id = {s.nhemployee.id:s.nhemployee for s in salaries}

    enrich_nh_employee_salaries(salaries, employee_by_id, year, month, year, month)

    set_loan_fields(employee_by_id.values())

    return render(request, 'Management/nhemployee_salaries.html', 
                              {'branch_list':branch_list, 'month': date(int(year), int(month), 1),
                               'filterForm':MonthForm(initial={'year':year,'month':month})},
                               )

@permission_required('Management.salaries_bank')
def salaries_bank(request):
    if request.method == 'POST':
        form = MonthForm(request.POST)
        if form.is_valid():
            month, year = form.cleaned_data['month'], form.cleaned_data['year']

            query = EmployeeSalaryBase.objects.select_related(
                'employeesalary__employee__employment_terms',
                'employeesalary__employee__account',
                'nhemployeesalary__nhemployee__employment_terms',
                'nhemployeesalary__nhemployee__account')

            if 'pdf' in request.POST:
                salary_ids = [key.replace('salary-','') for key in request.POST if key.startswith('salary-')]
                salaries = query.filter(pk__in = salary_ids)
            elif 'filter' in request.POST:
                salaries = query.filter(month=month, year=year)
                
            salaries.select_related(
                'employeesalary__employee__employment_terms',
                'employeesalary__employee__account',
                'nhemployeesalary__nhemployee__employment_terms',
                'nhemployeesalary__nhemployee__account')

            # replace base objects with derived
            salaries = [s.derived for s in salaries]

            # extract employees from salaries
            employees = [salary.get_employee() for salary in salaries]

            # create map
            employee_by_id = {e.id:e for e in employees}

            set_salary_base_fields(
                salaries, 
                employee_by_id,
                year, month, year, month)

            if 'pdf' in request.POST:
                writer = SalariesBankWriter(salaries, month, year)
                return build_and_return_pdf(writer)
        else:
            raise ValidationError
    else:
        now = datetime.now()
        month, year = now.month, now.year
        args = {'month':month, 'year':year}
        form = MonthForm(initial = args)
        salaries = EmployeeSalaryBase.objects.filter(**args)
    
    for salary in salaries:
        employee = salary.get_employee()
        if isinstance(employee, Employee):
            salary.division = u'נווה העיר'
        elif isinstance(employee, NHEmployee):
            salary.division = str(salary.nhemployeesalary.nhbranch)
    
    salaries = list(salaries)
    salaries.sort(key = lambda salary: salary.division)

    return render(request, 'Management/salaries_bank.html', 
                              {'salaries':salaries,'filterForm':form, 'month':datetime(year, month, 1)},
                               )  

class EmployeeSalaryUpdate(PermissionRequiredMixin, UpdateView):
    model = EmployeeSalary
    form_class = EmployeeSalaryForm
    template_name = 'Management/employee_salary_edit.html'
    permission_required = 'Management.change_employeesalary'

class NHEmployeeSalaryUpdate(PermissionRequiredMixin, UpdateView):
    model = NHEmployeeSalary
    form_class = NHEmployeeSalaryForm
    template_name = 'Management/nhemployee_salary_edit.html'
    permission_required = 'Management.change_nhemployeesalary'

def employee_salary_pdf(request, year, month):

    salaries = EmployeeSalary.objects \
        .select_related('employee__employment_terms__hire_type') \
        .nondeleted() \
        .filter(year = year, month= month) \
        .order_by('id')

    employee_by_id = {s.employee.id:s.employee for s in salaries}

    set_salary_base_fields(
        salaries, 
        employee_by_id, 
        year, month, year, month)

    set_salary_status_date(salaries)

    approved_salaries = [salary for salary in salaries if salary.approved_date]

    title = u'שכר עבודה למנהלי פרויקטים לחודש %s\%s' % (year, month)

    writer = EmployeeSalariesBookKeepingWriter(approved_salaries, title)

    return build_and_return_pdf(writer)

def employee_salary_calc(request, model, id):
    salary = model.objects.get(pk=id)
    
    year, month = salary.year, salary.month

    if model == EmployeeSalary:
        employee = salary.employee

        # enrish salary object
        set_employee_sales(
            [salary], 
            {employee.id:employee}, 
            {employee.id:list(employee.projects.all())},
            year, month, year, month)

    salary.calculate()
    salary.save()
    
    if model == EmployeeSalary:
        url = reverse('salary-list') + '?year=%s&month=%s' % (year, month)
    elif model == NHEmployeeSalary:
        url = reverse('nh-salary-list') + '?year=%s&month=%s' % (year, month)

    return HttpResponseRedirect(url)

@permission_required('Management.employee_salary_delete')
def employee_salary_delete(request, model, id):
    es = model.objects.get(pk = id)
    es.mark_deleted()
    es.save()
    
    if model == EmployeeSalary:
        url = reverse('salary-list') + '?year=%s&month=%s' % (es.year, es.month)
    elif model == NHEmployeeSalary:
        url = reverse('nh-salary-list') + '?year=%s&month=%s' % (es.year, es.month)

    return HttpResponseRedirect(url)

class EmployeeSalaryCommissionDetailView(LoginRequiredMixin, DetailView):
    model = EmployeeSalary
    context_object_name = 'salary'
    template_name = 'Management/employee_commission_details.html'

class EmployeeSalaryCheckDetailView(LoginRequiredMixin, DetailView):
    model = EmployeeSalary
    context_object_name = 'salary'
    template_name = 'Management/employee_salary_check_details.html'

class EmployeeSalaryTotalDetailView(LoginRequiredMixin, DetailView):
    model = EmployeeSalary
    context_object_name = 'salary'
    template_name = 'Management/employee_salary_total_details.html'
    

class NHEmployeeSalaryCommissionDetailView(LoginRequiredMixin, DetailView):
    model = NHEmployeeSalary
    context_object_name = 'salary'
    template_name = 'Management/nhemployee_commission_details.html'

class NHEmployeeSalaryCheckDetailView(LoginRequiredMixin, DetailView):
    model = NHEmployeeSalary
    context_object_name = 'salary'
    template_name = 'Management/employee_salary_check_details.html'

class NHEmployeeSalaryTotalDetailView(LoginRequiredMixin, DetailView):
    model = NHEmployeeSalary
    context_object_name = 'salary'
    template_name = 'Management/employee_salary_total_details.html'


@permission_required('Management.list_demand')
def demands_all(request):
    error = None
    if request.method == 'POST':
        houseForm = LocateHouseForm(request.POST)
        demandForm = LocateDemandForm(request.POST)
        if houseForm.is_valid():
            houses = House.objects.filter(building__project = houseForm.cleaned_data['project'], 
                                          building__num = houseForm.cleaned_data['building_num'],
                                          num = houseForm.cleaned_data['house_num'])
            if houses.count() > 0:
                return HttpResponseRedirect('/buildings/%s/house/%s/type1' % (houses[0].building.id, houses[0].id))
            else:
                error = u'לא נמצאה דירה מס %s בבניין מס %s בפרוייקט %s' % (houseForm.cleaned_data['house_num'],
                                                                           houseForm.cleaned_data['building_num'],
                                                                           houseForm.cleaned_data['project'])
        if demandForm.is_valid():
            demands = Demand.objects.filter(project = demandForm.cleaned_data['project'], 
                                            month = demandForm.cleaned_data['month'],
                                            year = demandForm.cleaned_data['year'])
            if demands.count()>0:
                return HttpResponseRedirect('/reports/project_month/%s/%s/%s' % (demands[0].project.id,
                                                                         demands[0].year, demands[0].month))
    
    total_mispaid, total_unpaid, total_nopayment, total_noinvoice, total_notyetpaid = 0,0,0,0,0
    amount_mispaid, amount_unpaid, amount_nopayment, amount_noinvoice, amount_notyetpaid = 0,0,0,0,0
    
    all_demands = Demand.objects \
        .all() \
        .prefetch_related('invoices__offset','payments','project__contacts') \
        .select_related('project__demand_contact','project__payment_contact') \
        .order_by('project_id','year','month')

    set_demand_diff_fields(all_demands)
    set_demand_invoice_payment_fields(all_demands)

    # group demands by project
    demand_groups = itertools.groupby(all_demands, lambda demand: demand.project)

    demands_by_project = {project:list(demand_iter) for (project, demand_iter) in demand_groups}

    projects = []
    
    for (project, demands) in demands_by_project.items():

        project.demands_mispaid = 0
        project.demands_unpaid = 0
        project.demands_nopayment = 0
        project.demands_noinvoice = 0

        for demand in demands:
            total_amount = demand.total_amount
            invoices_amount = demand.invoices_amount
            payments_amount = demand.payments_amount
            is_fully_paid = demand.is_fully_paid

            if is_fully_paid == False:
                if invoices_amount == None and payments_amount != None:
                    amount_noinvoice += total_amount
                    total_noinvoice += 1
                    project.demands_noinvoice += 1
                if invoices_amount != None and payments_amount == None:
                    amount_nopayment += total_amount
                    total_nopayment += 1
                    project.demands_nopayment += 1
                if invoices_amount != None and payments_amount != None and demand.diff_invoice_payment != 0:
                    amount_mispaid += total_amount
                    total_mispaid += 1
                    project.demands_mispaid += 1
                if invoices_amount == None and payments_amount == None:
                    amount_unpaid += total_amount
                    total_unpaid += 1
                    project.demands_unpaid += 1
                if demand.not_yet_paid == 1:
                    amount_notyetpaid += total_amount
                    total_notyetpaid += 1
                
        projects.append(project)

    context = { 
        'projects':projects, 'total_mispaid':total_mispaid, 'total_unpaid':total_unpaid,
        'total_nopayment':total_nopayment, 'total_noinvoice':total_noinvoice,'total_notyetpaid':total_notyetpaid,
        'amount_mispaid':amount_mispaid, 'amount_unpaid':amount_unpaid, 'amount_notyetpaid':amount_notyetpaid,
        'amount_nopayment':amount_nopayment, 'amount_noinvoice':amount_noinvoice,
        'houseForm':LocateHouseForm(), 'demandForm':LocateDemandForm(),
        'error':error }

    return render(request, 'Management/demands_all.html', context)

def employee_list_pdf(request):
    writer = EmployeeListWriter(
        employees = Employee.objects.active(),
        nhemployees = NHEmployee.objects.active())
    return build_and_return_pdf(writer)

class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'Management/employee_list.html'
    context_object_name = 'employee_list'

    def get_queryset(self):
        employees = Employee.objects \
            .active() \
            .select_related('employment_terms__hire_type','rank') \
            .prefetch_related('projects')

        return employees
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        
        context['nhbranch_list'] = NHBranch.objects.all()
        return context

class EmployeeArchiveListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'Management/employee_archive.html'
    context_object_name = 'employee'

    def get_queryset(self):
        employees = Employee.objects \
            .archive() \
            .select_related('employment_terms__hire_type','rank')           

        return employees
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        
        context['nhbranch_list'] = NHBranch.objects.all()
        return context

@permission_required('Management.nhmonth_season')
def nh_season_income(request):
    nhmonth_set, employees = NHMonth.objects.none(), []
    totals_notax = {'income':0, 'net_income':0, 'net_income_no_commission':0}
    totals = {'income':0, 'net_income':0, 'net_income_no_commission':0, 'sale_count':0}
    avg = {'signed_commission':0, 'actual_commission':0}
    avg_notax = {'income':0, 'net_income':0, 'net_income_no_commission':0}
    nhbranch = None
    
    if len(request.GET):
        form = NHBranchSeasonForm(request.GET)
        if form.is_valid():
            nhbranch = form.cleaned_data['nhbranch']
            if not request.user.has_perm('Management.nhbranch_' + str(nhbranch.id)):
                return HttpResponse('No Permission. Contact Elad.') 
            from_date = date(form.cleaned_data['from_year'], form.cleaned_data['from_month'], 1)
            to_year, to_month = form.cleaned_data['to_year'], form.cleaned_data['to_month']
            to_date = date(to_month == 12 and to_year + 1 or to_year, to_month == 12 and 1 or to_month + 1, 1)
        
            nhmonth_set = NHMonth.objects.range(from_date.year, from_date.month, to_date.year, to_date.month) \
                                         .filter(nhbranch = nhbranch).annotate(Count('nhsales'))
                
            query = NHBranchEmployee.objects.filter(start_date__lt = to_date, nhbranch = nhbranch) \
                                            .exclude(end_date__isnull=False, end_date__lt = from_date)
            employees = [x.nhemployee for x in query]
            
            for e in employees:
                e.season_total, e.season_total_notax, e.season_branch_income_notax = 0, 0, 0
                e.season_branch_income_buyers_notax, e.season_branch_income_sellers_notax = 0, 0
            for nhm in nhmonth_set:
                query = NHBranchEmployee.objects.filter(start_date__lt = to_date, nhbranch = nhbranch) \
                                                .exclude(end_date__isnull=False, end_date__lt = from_date)
                nhm.employees = [x.nhemployee for x in query]
                tax = Tax.objects.filter(date__lte=date(nhm.year, nhm.month,1)).latest().value / 100 + 1
                for e in nhm.employees:
                    e.month_total = 0
                for nhs in nhm.nhsales.all():
                    for nhss in nhs.nhsaleside_set.all():
                        for e in employees:
                            nhss.include_tax = True
                            e.season_total += nhss.get_employee_pay(e)
                            nhss.include_tax = False
                            e.season_total_notax += nhss.get_employee_pay(e)
                            if nhss.signing_advisor == e:
                                income_notax = nhss.income / tax
                                e.season_branch_income_notax += income_notax
                                if nhss.sale_type.id in [SaleType.SaleSeller, SaleType.RentRenter]:
                                    e.season_branch_income_sellers_notax += income_notax
                                elif nhss.sale_type.id in [SaleType.SaleBuyer, SaleType.RentRentee]:
                                    e.season_branch_income_buyers_notax += income_notax
                        for e in nhm.employees:
                            nhss.include_tax = True
                            e.month_total += nhss.get_employee_pay(e)
                nhm.include_tax = False
                totals_notax['income'] += nhm.total_income
                totals_notax['net_income'] += nhm.total_net_income
                totals_notax['net_income_no_commission'] += nhm.net_income_no_commission
                nhm.include_tax = True
                totals['income'] += nhm.total_income
                totals['net_income'] += nhm.total_net_income
                totals['net_income_no_commission'] += nhm.net_income_no_commission
                totals['sale_count'] += nhm.nhsales__count
                avg['signed_commission'] += nhm.avg_signed_commission
                avg['actual_commission'] += nhm.avg_actual_commission
            month_count = len(nhmonth_set)
            month_with_sales_count = len([nhm for nhm in nhmonth_set if nhm.nhsales__count > 0])
            if month_count > 0:
                avg['signed_commission'] /= month_with_sales_count
                avg['actual_commission'] /= month_with_sales_count
                avg_notax['income'] = totals_notax['income'] / month_count
                avg_notax['net_income'] = totals_notax['net_income'] / month_count
                avg_notax['net_income_no_commission'] = totals_notax['net_income_no_commission'] / month_count
               
            for e in employees:
                e.season_avg_notax = month_count and e.season_total_notax / month_count or 0
                total_net_income_notax = totals_notax['net_income']
                if total_net_income_notax:
                    e.season_branch_income_ratio_notax = e.season_branch_income_notax / total_net_income_notax * 100
                    e.season_branch_income_buyers_ratio_notax = e.season_branch_income_buyers_notax / total_net_income_notax * 100
                    e.season_branch_income_sellers_ratio_notax = e.season_branch_income_sellers_notax / total_net_income_notax * 100
    else:
        form = NHBranchSeasonForm()
    return render(request, 'Management/nh_season_income.html', 
                              { 'nhmonths':nhmonth_set, 'filterForm':form, 'employees':employees,
                               'totals':totals,'totals_notax':totals_notax,
                               'nhbranch':nhbranch, 'avg':avg, 'avg_notax':avg_notax },
                              )
    
class NHBranchCreate(PermissionRequiredMixin, CreateView):
    model = NHBranch
    fields = ('name','address','phone','mail','fax','url')
    template_name = 'Management/nhbranch_edit.html'
    permission_required = 'Management.add_nhbranch'

class NHBranchUpdate(PermissionRequiredMixin, UpdateView):
    model = NHBranch
    fields = ('name','address','phone','mail','fax','url')
    template_name = 'Management/nhbranch_edit.html'
    permission_required = 'Management.change_nhbranch'

def nhmonth_sales(request, nhbranch_id):
    if not request.user.has_perm('Management.nhbranch_' + str(nhbranch_id)):
        return HttpResponse('No Permission. Contact Elad.') 
    today = date.today()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    d = date(year, month, 1)
    if year and month:
        q = NHMonth.objects.filter(nhbranch__id = nhbranch_id, year=year, month=month)
    nhb = NHBranch.objects.get(pk=nhbranch_id)
    nhm = q.count() > 0 and q[0] or NHMonth(nhbranch = nhb, year = year, month = month)
    query = NHBranchEmployee.objects.filter(nhbranch = nhb).exclude(end_date__isnull=False, end_date__lt = d)
    employees = [x.nhemployee for x in query]
    for e in employees:
        e.month_total = 0
    for sale in nhm.nhsales.all():
        for nhss in sale.nhsaleside_set.all():
            for e in employees:
                e.month_total += nhss.get_employee_pay(e)
    form = MonthForm(initial={'year':nhm.year,'month':nhm.month})
    return render(request, 'Management/nhmonth_sales.html', 
                              { 'nhmonth':nhm, 'filterForm':form, 'employees':employees },
                              )

@permission_required('Management.change_nhmonth')
def nhmonth_close(request):
    today = date.today()
    nhbranch_id = int(request.GET.get('nhbranch'))
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    if not request.user.has_perm('Management.nhbranch_' + str(nhbranch_id)):
        return HttpResponse('No Permission. Contact Elad.')
    query = NHMonth.objects.filter(nhbranch__id = nhbranch_id, year = year, month = month)
    if query.count() == 1:
        nhm = query[0]
    elif query.count() == 0:
        nhbranch = NHBranch.objects.get(pk=nhbranch_id)
        nhm = NHMonth(nhbranch = nhbranch, year = year, month = month)
    nhm.close()
    return HttpResponseRedirect('/nhbranch/%s/sales?year=%s&month=%s' % (nhbranch_id, nhm.year, nhm.month))

@permission_required('Management.add_demand')
def demand_list(request):
    ds, unhandled_projects = [], []
    
    current = common.current_month()
    year, month = current.year, current.month
    
    if len(request.GET):
        form = MonthForm(request.GET)
        if form.is_valid():
            year, month = form.cleaned_data['year'], form.cleaned_data['month']
    else:
        form = MonthForm(initial={'year':year,'month':month})
        
    '''loop through all active projects and create demands for them if havent
    alredy created. if project has status other than Feed, it is handled''' 
    for project in Project.objects.active():
        demand, new = Demand.objects.get_or_create(project = project, year = year, month = month)
        ds.append(demand)

    set_demand_sale_fields(ds, year, month, year, month)
    set_demand_last_status(ds)
    set_demand_open_reminders(ds)

    # add un-handled projects
    for demand in ds:
        last_status = demand.last_status

        if last_status and last_status.type_id == DemandStatusType.Feed:
            unhandled_projects.append(demand.project)

    sales_count, expected_sales_count, sales_amount = 0,0,0

    for d in ds:
        sales_count += d.sales_count
        sales_amount += d.sales_amount
        expected_sales_count += d.sale_count
        
    context = { 
        'demands':ds, 
        'unhandled_projects':unhandled_projects, 
        'month':date(year, month, 1), 'filterForm':form, 'sales_count':sales_count ,
        'sales_amount':sales_amount, 'expected_sales_count':expected_sales_count }

    return render(request, 'Management/demand_list.html', context)

def employee_sales(request, id, year, month):
    salary = EmployeeSalary.objects.get(employee__id = id, year = year, month = month)

    employee = salary.employee

    set_employee_sales(
        [salary], 
        {id:employee}, 
        {id:list(employee.projects.all())},
        year, month, year, month)

    return render(request, 'Management/employee_sales.html', { 'es':salary })

@permission_required('Management.add_employeesalary')
def employee_refund(request, year, month):
    if request.method == 'POST':
        form = EmployeeSalaryRefundForm(request.POST)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            es = EmployeeSalary.objects.get_or_create(employee = cleaned_data['employee'], year = year, month = month)[0]
            es.refund = cleaned_data['refund']
            es.refund_type = cleaned_data['refund_type']
            es.save()
    else:
        form = EmployeeSalaryRefundForm()

    return render(request, 'Management/employee_salary_edit.html', 
                              { 'form':form, 'month':month, 'year':year },
                              )

@permission_required('Management.add_employeesalary')
def employee_remarks(request, year, month):
    if request.method == 'POST':
        form = EmployeeSalaryRemarksForm(request.POST)
        if form.is_valid():  
            cleaned_data = form.cleaned_data
            es = EmployeeSalary.objects.get_or_create(employee = cleaned_data['employee'], year = year, month = month)[0]
            es.remarks = cleaned_data['remarks']
            es.save()
    else:
        form = EmployeeSalaryRemarksForm()

    return render(request, 'Management/employee_salary_edit.html', 
                              { 'form':form, 'month':month, 'year':year },
                              )
    
def nhemployee_add(request):
    if request.method == 'POST':
        form = NHEmployeeForm(request.POST)
        if form.is_valid():
            nhemployee = form.save()
            nhbranch = form.cleaned_data['nhbranch']
            NHBranchEmployee.objects.create(nhemployee = nhemployee, nhbranch = nhbranch, start_date = nhemployee.work_start, 
                                            is_manager = False)
            return HttpResponseRedirect(nhemployee.get_absolute_url())
    else:
        form = NHEmployeeForm()
        
    return render(request, 'Management/nhemployee_form.html', {'form':form}, )

def nhemployee_sales(request, id, year, month):
    es = NHEmployeeSalary.objects.get(nhemployee__id = id, year = year, month = month)
    return render(request, 'Management/nhemployee_sales.html', 
                              { 'es':es },
                              )
    
@permission_required('Management.add_nhemployeesalary')
def nhemployee_refund(request, year, month):
    if request.method == 'POST':
        form = NHEmployeeSalaryRefundForm(request.POST)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            es = NHEmployeeSalary.objects.get_or_create(nhemployee = cleaned_data['nhemployee'], year = year, month = month)[0]
            es.refund = cleaned_data['refund']
            es.refund_type = cleaned_data['refund_type']
            es.save()
    else:
        form = NHEmployeeSalaryRefundForm()

    return render(request, 'Management/nhemployee_salary_edit.html', 
                              { 'form':form, 'month':month, 'year':year },
                              )

@permission_required('Management.add_nhemployeesalary')
def nhemployee_remarks(request, year, month):
    if request.method == 'POST':
        form = EmployeeSalaryRemarksForm(request.POST)
        if form.is_valid():            
            cleaned_data = form.cleaned_data
            es = NHEmployeeSalary.objects.get_or_create(nhemployee = cleaned_data['nhemployee'], year = year, month = month)[0]
            es.remarks = cleaned_data['remarks']
            es.save()
    else:
        form = NHEmployeeSalaryRemarksForm()

    return render(request, 'Management/nhemployee_salary_edit.html', 
                              { 'form':form, 'month':month, 'year':year },
                              )

@permission_required('Management.change_demand')
def demand_edit(request, object_id):
    demand = Demand.objects \
        .prefetch_related('diffs','invoices','payments') \
        .select_related('project') \
        .get(pk = object_id)

    year, month = demand.year, demand.month

    set_demand_sale_fields([demand], year, month, year, month)
    set_demand_diff_fields([demand])

    months = None

    if request.method == 'POST':
        form = DemandForm(request.POST, instance = demand)
        if form.is_valid():
            form.save()
    else:
        form = DemandForm(instance = demand)
        months = MonthForm()
        
    # set sales_commission_base: summary of pc_base for each sale
    demand.sales_commission_base = sum(map(lambda sale: sale.pc_base_worth, demand.sales_list))

    context = { 
        'form':form, 
        'demand':demand, 
        'sales':demand.sales_list, 
        'months':months 
    }

    return render(request, 'Management/demand_edit.html', context)
    
def demand_sales_export(request, id):
    demand = Demand.objects \
        .prefetch_related('diffs','invoices','payments') \
        .select_related('project') \
        .get(pk=id)

    year, month = demand.year, demand.month

    set_demand_sale_fields([demand], year, month, year, month)

    sales = demand.sales_list
    commissions = demand.project.commissions.get()

    discount_cols = sales[0].discount != None
    bonus_cols = commissions.b_discount_save_precentage != None
    zilber_cols = commissions.c_zilber != None

    columns = [
        ExcelColumn('מזהה מכירה'),
        ExcelColumn('שם הרוכשים', width=20),
        ExcelColumn('בניין'),
        ExcelColumn('דירה'),
        ExcelColumn('ת. הרשמה', width=15),
        ExcelColumn('ת. מכירה', width=15),
        ExcelColumn('מחיר חוזה', 'currency', showSum=True, width=20),
        ExcelColumn('מחיר חוזה לעמלה', 'currency', showSum=True, width=20),
    ]

    if discount_cols:
        columns += [
            ExcelColumn('מחיר כולל מע"מ', 'currency', showSum=True),
            ExcelColumn('% הנחה ניתן', 'percent'),
            ExcelColumn('% הנחה מותר', 'percent'),
        ]
        
    columns += [
        ExcelColumn('% עמלת בסיס', 'percent', width=15),
        ExcelColumn('שווי עמלת בסיס', 'currency', showSum=True, width=20),
    ]
    
    if bonus_cols or zilber_cols:
        columns += [
            ExcelColumn('% בונוס חיסכון', 'percent'),
            ExcelColumn('שווי בונוס חיסכון', 'currency', showSum=True),
        ]

    columns += [
        ExcelColumn('% עמלה סופי', 'percent', width=15),
        ExcelColumn('שווי עמלה סופי', 'currency', showSum=True, width=20),
    ]

    rows = []

    for counter, sale in enumerate(sales, 1):
        signup = sale.house.get_signup()

        row = [
            '{demand_id}-{counter}'.format(demand_id=id, counter=counter),
            sale.clients,
            sale.house.building.num,
            sale.house.num,
            signup.date if signup != None else '',
            sale.sale_date,
            sale.price,
            sale.price_final
        ]

        if discount_cols:
            row += [
                sale.price_taxed,
                sale.discount,
                sale.allowed_discount
            ]
        
        row += [
            sale.pc_base,
            sale.pc_base_worth
        ]

        if bonus_cols:
            row += [
                sale.pb_dsp,
                sale.pb_dsp_worth
            ]
        elif zilber_cols:
            row += [
                '',
                sale.zdb
            ]

        row += [
            sale.c_final,
            sale.c_final_worth
        ]

        rows.append(row)
    
    title = u'פירוט מכירות - {project} לחודש {month}/{year}'.format(
        project=demand.project.name, month=month, year=year)

    return ExcelGenerator().generate(title, columns, rows)

@permission_required('Management.change_demand')
def demand_close(request, id):
    demand = Demand.objects.get(pk=id)
    
    year, month = demand.year, demand.month

    set_demand_sale_fields([demand], year, month, year, month)

    if request.method == 'POST':
        demand.close()
        demand.save()

    return render(request, 'Management/demand_close.html', 
        { 'demand':demand })

@permission_required('Management.change_demand')
def demand_zero(request, id):
    d = Demand.objects.get(pk=id)
    d.close()
    return redirect('demand-list')

@permission_required('Management.send_mail')
def send_mail(request):
    if request.method == 'POST':
        form = MailForm(request.POST, request.FILES)
        if form.is_valid():
            field_names = ['attachment1','attachment2','attachment3','attachment4','attachment5']
            attachments = [request.FILES[field_name] for field_name in field_names if form.cleaned_data[field_name]]
            
            mail(to = form.cleaned_data['to'], cc = form.cleaned_data['Cc'], bcc = form.cleaned_data['Bcc'],
                 subject = form.cleaned_data['subject'], contents = form.cleaned_data['contents'], attachments = attachments)
    else:
        form = MailForm()
        
    return render(request, 'Management/send_mail.html',  { 'form':form }, )

def demand_send_mail(demand, addr):
    filename = common.generate_unique_media_filename('pdf')
    MonthDemandWriter(demand, to_mail=True).build(filename)
    mail(to = addr, subject = u'עמלה לפרויקט %s לחודש %s/%s' % (demand.project, demand.month, demand.year), attachments = [filename])
    demand.send()

@permission_required('Management.change_demand')
def demands_send(request):
    current = common.current_month()
    y = int(request.GET.get('year', current.year))
    m = int(request.GET.get('month', current.month))
    form = MonthForm(initial={'year':y,'month':m})
    month = datetime(y,m,1)
    
    demands = Demand.objects \
        .filter(year = y, month = m) \
        .select_related('project__demand_contact')

    set_demand_sale_fields(demands, y, m, y, m)
    set_demand_diff_fields(demands)
    set_demand_last_status(demands)

    forms=[]
    if request.method == 'POST':
        error = False
        for demand in demands:
            f = DemandSendForm(request.POST, instance=demand, prefix = str(demand.id))
            if f.is_valid():
                is_finished, by_mail, by_fax, mail = f.cleaned_data['is_finished'], f.cleaned_data['by_mail'], f.cleaned_data['by_fax'], f.cleaned_data['mail']
                if is_finished:
                    demand.finish()
                if by_mail and demand.sales_count > 0:
                    if mail:
                        demand_send_mail(demand, mail)
                    else:
                        error = u'לפרויקט %s לא הוגדר מייל לשליחת דרישות' % demand.project
                if by_fax:
                    pass
            forms.append(f)
        if error:
            return render(request, 'Management/error.html', {'error': error}, )
        else:
            return HttpResponseRedirect(reverse('demand-old'))
    else:
        for demand in demands:
            contact = demand.project.demand_contact
            
            initial = { 'mail':contact.mail, 'fax':contact.fax } if contact else {}

            f = DemandSendForm(instance=demand, prefix=str(demand.id), initial = initial)
            
            forms.append(f)
            
    return render(request, 'Management/demands_send.html', 
        { 'forms':forms,'filterForm':form, 'month':month })

@permission_required('Management.change_demand')
def demand_closeall(request):
    for d in Demand.objects.current():
        if d.statuses.latest().type.id == DemandStatusType.Feed:
            d.close()
    return HttpResponseRedirect('/demands')
        
@permission_required('Management.delete_sale')
def demand_sale_del(request, demand_id, id):
    sale = Sale.objects.get(pk=id)
    if sale.demand.statuses.latest().type.id != DemandStatusType.Sent:
        sale.delete()
        return HttpResponseRedirect('../../../')
    else:
        sc = SaleCancel(sale = sale, date = date.today(), deduct_from_demand=True)
        sc.save()
        return HttpResponseRedirect('/salecancel/%s' % sc.id)

class SalePriceModUpdate(PermissionRequiredMixin, UpdateView):
    model = SalePriceMod
    form_class = SalePriceModForm
    template_name = 'Management/sale_mod_edit.html'
    permission_required = 'Management.change_salepricemod'
    
class SaleHouseModUpdate(PermissionRequiredMixin, UpdateView):
    model = SaleHouseMod
    form_class = SaleHouseModForm
    template_name = 'Management/sale_mod_edit.html'
    permission_required = 'Management.change_salehousemod'

class SaleCancelUpdate(PermissionRequiredMixin, UpdateView):
    model = SaleCancel
    fields = ('deduct_from_demand','remarks')
    template_name = 'Management/sale_mod_edit.html'
    permission_required = 'Management.change_salecancel'

def salepaymod_edit(request, model, object_id):
    from django.forms.models import modelform_factory

    object = model.objects.select_related('sale__demand__project', 'sale__house').get(pk = object_id)
    form_class = modelform_factory(model, fields=('to_month','to_year','employee_pay_month','employee_pay_year','remarks'))
    
    if request.method == 'POST':
        form = form_class(request.POST, instance = object)
        if form.is_valid():
            demands_to_calc = []
            salaries_to_calc = []
            to_year, to_month = form.cleaned_data['to_year'], form.cleaned_data['to_month']
            employee_pay_year, employee_pay_month = form.cleaned_data['employee_pay_year'], form.cleaned_data['employee_pay_month']
            project = object.sale.demand.project
            
            # need to re-calc both origin and destination demand (if changed)
            if object.to_year != to_year or object.to_month != to_month:
                q = Q(year = object.to_year, month = object.to_month) | Q(year = to_year, month = to_month)
                demands_to_calc.extend(project.demands.filter(q))
                
            # need to calc origin and destination salaries for all project employees
            if object.employee_pay_year != employee_pay_year or object.employee_pay_month != employee_pay_month:
                q = Q(year = object.employee_pay_year, month = object.employee_pay_month) | Q(year = employee_pay_year, month = employee_pay_month)
                salaries = EmployeeSalary.objects.nondeleted().filter(q, employee__in = project.employees.all())
                salaries_to_calc.extend(salaries)
            
            # enrich demands
            set_demand_diff_fields(demands_to_calc)

            for demand in demands_to_calc:
                year, month = demand.year, demand.month
                set_demand_sale_fields([demand], year, month, year, month)
                demand.calc_sales_commission()
            
            for salary in salaries_to_calc:
                salary.calculate()
                salary.save()
    else:
        form = form_class(instance = object)
        
    return render(request, 'Management/sale_mod_edit.html', {'form':form}, )
        

@permission_required('Management.reject_sale')
def demand_sale_reject(request, demand_id, id):
    sale = Sale.objects.get(pk=id)
    y,m = sale.contractor_pay_year, sale.contractor_pay_month
    demands_to_calc = []
    
    try:
        sr = sale.salereject
    except SaleReject.DoesNotExist:
        sr = SaleReject(sale = sale, employee_pay_month = sale.employee_pay_month, employee_pay_year = sale.employee_pay_year)
    
    to_year, to_month = m==12 and y+1 or y, m==12 and 1 or m+1
    
    if to_year != sr.to_year or to_month != sr.to_month:
        project = sale.demand.project
        # need to recalc both origin and destination demands
        q = Q(year = y, month = m) | Q(year = to_year, month = to_month)
        demands_to_calc.extend(project.demands.filter(q))
        
    sr.to_year, sr.to_month = to_year, to_month
    sr.save()
    
    # enrich demands
    set_demand_diff_fields(demands_to_calc)

    for demand in demands_to_calc:
        year, month = demand.year, demand.month
        set_demand_sale_fields([demand], year, month, year, month)
        demand.calc_sales_commission()
    
    return HttpResponseRedirect('/salereject/%s' % sr.id)

@permission_required('Management.pre_sale')
def demand_sale_pre(request, demand_id, id):
    sale = Sale.objects.get(pk=id)
    y,m = request.GET.get('year'), request.GET.get('month')
    demands_to_calc = []
    
    try:
        sr = sale.salepre
    except SalePre.DoesNotExist:
        sr = SalePre(sale = sale, employee_pay_month = sale.employee_pay_month, employee_pay_year = sale.employee_pay_year)
    
    to_year, to_month = y is None and sale.contractor_pay_year or y, m is None and sale.contractor_pay_month or m
    
    if to_year != sr.to_year or to_month != sr.to_month:
        project = sale.demand.project
        # need to recalc both origin and destination demands
        q = Q(year = y, month = m) | Q(year = to_year, month = to_month)
        demands_to_calc.extend(project.demands.filter(q))
        
    sr.to_year, sr.to_month = to_year, to_month
    sr.save()
    
    # enrich demands
    set_demand_diff_fields(demands_to_calc)

    for demand in demands_to_calc:
        year, month = demand.year, demand.month
        set_demand_sale_fields([demand], year, month, year, month)
        demand.calc_sales_commission()
    
    return HttpResponseRedirect('/salepre/%s' % sr.id)

@permission_required('Management.cancel_sale')
def demand_sale_cancel(request, demand_id, id):
    sale = Sale.objects.get(pk=id)
    try:
        sc = sale.salecancel
    except SaleCancel.DoesNotExist:
        sc = SaleCancel(sale = sale)
    sc.save()
    
    sale.commission_include = False
    sale.save()
    
    demand = sale.demand
    year, month = demand.year, demand.month

    # enrich demand
    set_demand_sale_fields([demand], year, month, year, month)
    set_demand_diff_fields([demand])
    
    #re-calculate the entire demand
    sale.demand.calc_sales_commission()
    
    return HttpResponseRedirect('/salecancel/%s' % sc.id)

@permission_required('Management.add_invoice')
def invoice_add(request, initial=None):
    if request.method == 'POST':
        form = DemandInvoiceForm(request.POST)
        if form.is_valid():
            form.save()
            if 'addanother' in request.POST:
                form = DemandInvoiceForm(initial=initial)
            if 'addpayment' in request.POST:
                return HttpResponseRedirect('/payments/add')
    else:
        form = DemandInvoiceForm(initial=initial)

    demand = initial['demand']

    context = {'form':form, 'sales':demand.sales_list, 'demand': demand}

    return render(request, 'Management/invoice_edit.html', context)

class InvoiceUpdate(PermissionRequiredMixin, UpdateView):
    model = Invoice
    form_class = InvoiceForm
    template_name = 'Management/invoice_edit.html'
    permission_required = 'Management.change_invoice'

@permission_required('Management.add_invoice')
def demand_invoice_add(request, id):
    demand = Demand.objects \
        .select_related('project') \
        .get(pk=id)

    year, month = demand.year, demand.month

    set_demand_sale_fields([demand], year, month, year, month)

    return invoice_add(request, {'project':demand.project_id, 'month':month, 'year':year, 'demand': demand})

@permission_required('Management.demand_invoices')
def demand_invoice_list(request):
    invoices = None
    if len(request.GET):
        form = ProjectSeasonForm(request.GET)
        if form.is_valid():
            project = form.cleaned_data['project']
            from_date = date(form.cleaned_data['from_year'], form.cleaned_data['from_month'], 1)
            to_date = date(form.cleaned_data['to_year'], form.cleaned_data['to_month'], 1)
            to_date = date(to_date.month == 12 and to_date.year + 1 or to_date.year, to_date.month == 12 and 1 or to_date.month + 1, 1)

            query = Invoice.objects.reverse().select_related()

            invoices = []
            for invoice in query:
                if project:
                    invoice_demands = invoice.demands.filter(project = project)
                else:
                    invoice_demands = invoice.demands.all()
                    
                if not invoice_demands.count():
                    continue
                
                for demand in invoice_demands:
                    demand_date = date(demand.year, demand.month, 1)
                    if demand_date >= from_date and demand_date < to_date:
                        invoices.append(invoice)
                        break

            paginator = django.core.paginator.Paginator(invoices, 25) 
        
            try:
                page = int(request.GET.get('page', '1'))
            except ValueError:
                page = 1
        
            try:
                invoices = paginator.page(page)
            except (paginator.EmptyPage, paginator.InvalidPage):
                invoices = paginator.page(paginator.num_pages)
    else:
        form = ProjectSeasonForm()

    return render(request, 'Management/demand_invoice_list.html', {'page': invoices,'filterForm':form})    
 
 
@permission_required('Management.demand_payments')
def demand_payment_list(request):
    payments = None
    if len(request.GET):
        form = ProjectSeasonForm(request.GET)
        if form.is_valid():
            project = form.cleaned_data['project']
            from_date = date(form.cleaned_data['from_year'], form.cleaned_data['from_month'], 1)
            to_date = date(form.cleaned_data['to_year'], form.cleaned_data['to_month'], 1)
            to_date = date(to_date.month == 12 and to_date.year + 1 or to_date.year, to_date.month == 12 and 1 or to_date.month + 1, 1)

            query = Payment.objects.reverse().select_related()
            
            payments = []
            for payment in query:
                if project:
                    payment_demands = payment.demands.filter(project = project)
                else:
                    payment_demands = payment.demands.all()
                    
                if not payment_demands.count():
                    continue
                
                for demand in payment_demands:
                    demand_date = date(demand.year, demand.month, 1)
                    if demand_date >= from_date and demand_date < to_date:
                        payments.append(payment)
                        break
            
            paginator = django.core.paginator.Paginator(payments, 25) 
        
            try:
                page = int(request.GET.get('page', '1'))
            except ValueError:
                page = 1
        
            try:
                payments = paginator.page(page)
            except (paginator.EmptyPage, paginator.InvalidPage):
                payments = paginator.page(paginator.num_pages)
    else:
        form = ProjectSeasonForm()

    return render(request, 'Management/demand_payment_list.html', {'page': payments,'filterForm':form})    
   
@permission_required('Management.add_invoice')
def project_invoice_add(request, id):
    return invoice_add(request, {'project':id})

@permission_required('Management.delete_invoice')
def invoice_del(request, id):
    i = Invoice.objects.get(pk=id)
    demand_id = i.demands.all()[0].id
    i.delete()
    return HttpResponseRedirect('/demands/%s' % demand_id)

@permission_required('Management.add_invoiceoffset')
def invoice_offset(request, id=None):
    if request.method == 'POST':
        form = InvoiceOffsetForm(request.POST)
        if form.is_valid():
            invoice_num = form.cleaned_data['invoice_num']
            invoice = Invoice.objects.get(num = invoice_num) 
            invoice.offset = form.save()
            invoice.save()
    else:
        if id:
            i = Invoice.objects.get(pk=id)
            invoice_num = i.num
            offset = i.offset or InvoiceOffset()
        else:
            offset = InvoiceOffset()
            invoice_num = 0
        form = InvoiceOffsetForm(instance = offset, initial={'invoice_num':invoice_num})
    
    return render(request, 'Management/object_edit.html', {'form': form, 'title':u'זיכוי חשבונית'})    

class InvoiceOffsetUpdate(PermissionRequiredMixin, UpdateView):
    model = InvoiceOffset
    fields = ('date','amount','reason','remarks')
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_invoiceoffset'

class LoanPayUpdate(PermissionRequiredMixin, UpdateView):
    model = LoanPay
    form_class = LoanPayForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_loanpay'

@permission_required('Management.delete_invoiceoffset')
def invoice_offset_del(request, id):
    io = InvoiceOffset.objects.get(pk=id)
    demand_id = io.invoice.demands.all()[0].id
    #unlink invoice from the offset
    invoice = io.invoice
    invoice.offset = None
    invoice.save()
    InvoiceOffset.objects.get(pk=id).delete()
    return HttpResponseRedirect('/demands/%s' % demand_id)

@permission_required('Management.add_payment')
def split_payment_add(request):
    DemandFormset = formset_factory(SplitPaymentDemandForm, extra=5)
    error = ''
    if request.method == 'POST':
        spf = SplitPaymentForm(request.POST)
        spdForms = DemandFormset(request.POST)
        if spf.is_valid() and spdForms.is_valid():
            for form in spdForms.forms:
                if not form.is_valid() or not 'amount' in form.cleaned_data:
                    continue
                p = Payment()
                for attr, value in spf.cleaned_data.items():
                    setattr(p, attr, value)
                p.amount = form.cleaned_data['amount']
                p.save()
                project, year, month = form.cleaned_data['project'], form.cleaned_data['year'], form.cleaned_data['month']
                q = Demand.objects.filter(project = project, year = year, month = month)
                if q.count() == 1:
                    q[0].payments.add(p)
                else:
                    error += '\r\n' + u'לא קיימת דרישה לפרוייקט %s לחודש %s\%s' % (project, year, month)
            if error == '':
                return HttpResponseRedirect('/demandpayments')
    else:
        spf = SplitPaymentForm()
        spdForms = DemandFormset()
        
    return render(request, 'Management/split_payment_add.html', 
                              { 'spf':spf, 'spdForms':spdForms, 'error':error }, )

@permission_required('Management.add_payment')
def payment_add(request, initial=None):
    if request.method == 'POST':
        form = DemandPaymentForm(request.POST)
        if form.is_valid():
            form.save()
            if 'addanother' in request.POST:
                form = DemandPaymentForm(initial=initial)
            if 'addinvoice' in request.POST:
                return HttpResponseRedirect('/invoices/add')
    else:
        form = DemandPaymentForm(initial=initial)
    return render(request, 'Management/payment_edit.html', 
                              { 'form':form }, )

class PaymentUpdate(PermissionRequiredMixin, UpdateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'Management/payment_edit.html'
    permission_required = 'Management.change_payment'

@permission_required('Management.change_payment')
def demand_payment_edit(request, id):
    payment = Payment.objects.get(pk = id)
    if request.method == 'POST':
        form = DemandPaymentForm(request.POST, instance = payment)
        if form.is_valid():
            form.save()
            if 'addanother' in request.POST:
                try:
                    demand = payment.demands.all()[0]
                    return HttpResponseRedirect('/demands/%s/payment/add' % demand.id)
                except KeyError:
                    return HttpResponseRedirect(reverse(payment_add))
            if 'addinvoice' in request.POST:
                return HttpResponseRedirect('/invoices/add')
    else:
        form = DemandPaymentForm(instance = payment)
    return render(request, 'Management/payment_edit.html', 
                              { 'form':form }, )

@permission_required('Management.change_invoice')
def demand_invoice_edit(request, id):
    invoice = Invoice.objects.get(pk = id)
    if request.method == 'POST':
        form = DemandInvoiceForm(request.POST, instance = invoice)
        if form.is_valid():
            form.save()
            if 'addanother' in request.POST:
                try:
                    demand = invoice.demands.all()[0]
                    return HttpResponseRedirect('/demands/%s/invoice/add' % demand.id)
                except KeyError:
                    return HttpResponseRedirect(reverse(invoice_add))
            if 'addpayment' in request.POST:
                return HttpResponseRedirect('/payments/add')
    else:
        form = DemandInvoiceForm(instance = invoice)
    return render(request, 'Management/invoice_edit.html', 
                              { 'form':form }, )

def payment_details(request, project, year, month):
    try:
        d = Demand.objects.get(project = project, year = year, month = month)
        return render(request, 'Management/demand_payment_details.html', 
                                  { 'payments':d.payments.all()}, )
    except Demand.DoesNotExist:
        return HttpResponse('')
    
def invoice_details(request, project, year, month):
    try:
        d = Demand.objects.get(project = project, year = year, month = month)
        return render(request, 'Management/demand_invoice_details.html', 
                                  { 'invoices':d.invoices.all()}, )
    except Demand.DoesNotExist:
        return HttpResponse('')
    
def demand_details(request, project, year, month):
    try:
        demand = Demand.objects.get(project=project, year=year, month=month)

        set_demand_diff_fields([demand])

        return render(request, 'Management/demand_details.html', { 'demand':demand})
    except Demand.DoesNotExist:
        return HttpResponse('')

@permission_required('Management.add_demanddiff')
def demand_adddiff(request, object_id, type = None):
    demand = Demand.objects.get(pk=object_id)
    if request.method == 'POST':
        form = DemandDiffForm(request.POST)
        if form.is_valid():
            form.instance.demand = demand
            diff = form.save()
            if 'addanother' in request.POST:
                return HttpResponseRedirect(reverse(demand_adddiff, args=[object_id]))
            return HttpResponseRedirect(diff.get_absolute_url())
    else:
        form = DemandDiffForm(initial={'type':type})
    return render(request, 'Management/object_edit.html', 
                              { 'form':form }, )

@permission_required('Management.demand_force_fully_paid')
def demand_force_fully_paid(request, id):
    demand = Demand.objects.get(pk=id)
    demand.force_fully_paid = True
    demand.save()
    return HttpResponse('ok')

@permission_required('Management.demand_not_yet_paid')
def demand_not_yet_paid(request, id):
    demand = Demand.objects.get(pk=id)
    demand.not_yet_paid = demand.not_yet_paid == 0 and True or False
    demand.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

@permission_required('Management.add_demanddiff')
def demand_adddiff_adjust(request, object_id):
    return demand_adddiff(request, object_id, u'התאמה')

class DemandDiffUpdate(PermissionRequiredMixin, UpdateView):
    model = DemandDiff
    form_class = DemandDiffForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_demanddiff'

@permission_required('Management.delete_demanddiff')
def demanddiff_del(request, pk):
    diff = DemandDiff.objects.get(pk=pk)
    url = diff.demand.get_absolute_url()
    diff.delete()
    return HttpResponseRedirect(url)

@permission_required('Management.add_payment')
def demand_payment_add(request, id):
    demand = Demand.objects.get(pk=id)
    return payment_add(request, {'project':demand.project_id, 'month':demand.month, 'year':demand.year})
    
@permission_required('Management.add_payment')
def project_payment_add(request, id):    
    return payment_add(request, {'project':id})

@permission_required('Management.add_payment')
def nhsaleside_payment_add(request, object_id):
    nhs = NHSaleSide.objects.get(pk=object_id)
    if not request.user.has_perm('Management.nhbranch_' + str(nhs.nhsale.nhmonth.nhbranch.id)):
        return HttpResponse('No Permission. Contact Elad.') 
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            p = form.save()
            nhs.payments.add(p)
            if 'addanother' in request.POST:
                form = PaymentForm()
    else:
        form = PaymentForm()
    return render(request, 'Management/object_edit.html', 
                              { 'form':form }, )
    
@permission_required('Management.add_invoice')
def nhsaleside_invoice_add(request, object_id):
    nhs = NHSaleSide.objects.get(pk=object_id)
    if not request.user.has_perm('Management.nhbranch_' + str(nhs.nhsale.nhmonth.nhbranch.id)):
        return HttpResponse('No Permission. Contact Elad.') 
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            i = form.save()
            nhs.invoices.add(i)
    else:
        form = InvoiceForm()
    return render(request, 'Management/object_edit.html', 
                              { 'form':form }, )

 
@permission_required('Management.delete_payment')
def payment_del(request, id):
    p = Payment.objects.get(pk=id)
    if p.demands.count() == 1:
        next = '/demands/%s' % p.demands.all()[0].id
    elif i.nhsaleside_set.count() == 1:
        next = '/nhsale/%s' % p.nhsaleside_set.all()[0].nhsale.id
    p.delete()
    return HttpResponseRedirect(next)

class ProjectListView(LoginRequiredMixin, ListView):
    model = Project

    def get_queryset(self):
        projects = Project.objects \
            .select_related('details','demand_contact','payment_contact') \
            .prefetch_related('employees','contacts') \
            .filter(end_date__isnull = True)

        set_project_buildings(projects)

        return projects

@permission_required('Management.project_list_pdf')
def project_list_pdf(request):
    writer = ProjectListWriter(projects = Project.objects.active())
    return build_and_return_pdf(writer)

def set_project_buildings(projects):
    # load non-deleted buildings for all projects
    project_ids = [project.id for project in projects]

    buildings = Building.objects \
        .filter(is_deleted=False, project_id__in=project_ids) \
        .order_by('project_id')

    # construct a map of project_id -> buildings
    building_groups = itertools.groupby(buildings, lambda building: building.project_id)

    buildings_map = {project_id: list(buildings_iter) for (project_id, buildings_iter) in building_groups}

    for project in projects:
        project.non_deleted_buildings = buildings_map.get(project.id, [])

class ProjectArchiveListView(LoginRequiredMixin, ListView):
    model = Project
    template_name = 'Management/project_archive.html'

    def get_queryset(self):
        projects = Project.objects \
            .select_related('details','demand_contact','payment_contact') \
            .prefetch_related('employees','contacts') \
            .filter(end_date__isnull = False)

        set_project_buildings(projects)

        return projects

class NHSaleDetailView(LoginRequiredMixin, DetailView):
    model = NHSale
    context_object_name = 'nhs'
    template_name = 'Management/nhsale_edit.html'

class NHSaleUpdate(PermissionRequiredMixin, UpdateView):
    model = NHSale
    form_class = NHSaleForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_nhsale'

    
class NHSaleSideUpdate(PermissionRequiredMixin, UpdateView):
    model = NHSaleSide
    form_class = NHSaleSideForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_nhsaleside'

@permission_required('Management.change_nhsale')
def nhsale_edit(request, object_id):
    nhs = NHSale.objects.get(pk=object_id)
    if not request.user.has_perm('Management.nhbranch_' + str(nhs.nhmonth.nhbranch.id)):
        return HttpResponse('No Permission. Contact Elad.') 
    return render(request, 'Management/nhsale_edit.html',
                              {'nhs': nhs}, 
                              )

@permission_required('Management.nhsale_move_nhmonth')
def nhsale_move_nhmonth(request, object_id):
    nhsale = NHSale.objects.get(pk=object_id)
    if request.method == 'POST':
        form = NHMonthForm(request.POST)
        if form.is_valid():
            kwargs = form.cleaned_data
            nhmonth, new = NHMonth.objects.get_or_create(**kwargs)
            nhsale.nhmonth = nhmonth
            nhsale.save()
    else:
        nhmonth = nhsale.nhmonth
        form = NHMonthForm(initial = {'year':nhmonth.year, 'month':nhmonth.month, 'nhbranch':nhmonth.nhbranch.id})
        
    return render(request, 'Management/object_edit.html',
                              {'form': form, 'title':u'העברת עסקה מס ' + str(nhsale.num)}, 
                              )

@permission_required('Management.add_nhsale')
def nhsale_add(request, branch_id):
    from django.forms.models import modelformset_factory

    if not request.user.has_perm('Management.nhbranch_' + str(branch_id)):
        return HttpResponse('No Permission. Contact Elad.') 
    
    PaymentFormset = modelformset_factory(Payment, PaymentForm, extra=5)

    if request.method=='POST':
        monthForm = NHMonthForm(request.POST, prefix='month')
        saleForm = NHSaleForm(request.POST, prefix='sale')
        side1Form = NHSaleSideForm(request.POST, prefix='side1')
        side2Form = NHSaleSideForm(request.POST, prefix='side2')
        invoice1Form = InvoiceForm(request.POST, prefix='invoice1')
        payment1Forms = PaymentFormset(request.POST, prefix='payments1')
        invoice2Form = InvoiceForm(request.POST, prefix='invoice2')
        payment2Forms = PaymentFormset(request.POST, prefix='payments2')
        
        if monthForm.is_valid() and saleForm.is_valid() and side1Form.is_valid() and side2Form.is_valid():
            kwargs = monthForm.cleaned_data
            nhmonth, new = NHMonth.objects.get_or_create(**kwargs)
            saleForm.instance.nhmonth = nhmonth
            nhsale = saleForm.save()
            side1Form.instance.nhsale = side2Form.instance.nhsale = nhsale
            side1, side2 = side1Form.save(), side2Form.save()
            error = False
            if invoice1Form.has_changed() and invoice1Form.is_valid():
                side1.invoices.add(invoice1Form.save())
            else:
                error = invoice1Form.has_changed()
            if payment1Forms.is_valid():
                for p in payment1Forms.save():
                    side1.payments.add(p)
            else:
                error = True
            if invoice2Form.has_changed() and invoice2Form.is_valid():
                side2.invoices.add(invoice2Form.save())
            else:
                error = invoice2Form.has_changed()
            if payment2Forms.is_valid():
                for p in payment2Forms.save():
                    side2.payments.add(p)
            else:
                error = True
            if not error:
                if 'addanother' in request.POST:
                    return HttpResponseRedirect('add')
                elif 'tomonth' in request.POST:
                    return HttpResponseRedirect('/nhbranch/%s/sales' % nhsale.nhmonth.nhbranch.id)
    else:
        branch = NHBranch.objects.get(pk=branch_id)
        employee_base_query = EmployeeBase.objects.active()
        nhemployees_query = NHEmployee.objects.filter(nhbranchemployee__nhbranch = branch, work_end = None)
        
        saleForm = NHSaleForm(prefix='sale')
        monthForm = NHMonthForm(prefix='month')
        monthForm.fields['nhbranch'].initial = branch_id
        side1Form = NHSaleSideForm(prefix='side1')
        side1Form.fields['employee1'].queryset = nhemployees_query
        side1Form.fields['employee2'].queryset = nhemployees_query
        side1Form.fields['signing_advisor'].queryset = nhemployees_query
        side1Form.fields['director'].queryset = employee_base_query
        side2Form = NHSaleSideForm(prefix='side2')
        side2Form.fields['employee1'].queryset = nhemployees_query
        side2Form.fields['employee2'].queryset = nhemployees_query
        side2Form.fields['signing_advisor'].queryset = nhemployees_query
        side2Form.fields['director'].queryset = employee_base_query
        invoice1Form = InvoiceForm(prefix='invoice1')
        payment1Forms = PaymentFormset(prefix='payments1', queryset=Payment.objects.none())
        invoice2Form = InvoiceForm(prefix='invoice2')
        payment2Forms = PaymentFormset(prefix='payments2', queryset=Payment.objects.none())
        
    return render(request, 'Management/nhsale_add.html',
                              {'monthForm':monthForm, 'saleForm':saleForm, 
                               'side1form':side1Form, 'side2form':side2Form, 
                               'invoice1Form':invoice1Form, 'payment1Forms':payment1Forms, 
                               'invoice2Form':invoice2Form, 'payment2Forms':payment2Forms}, 
                              )

@permission_required('Management.change_pricelist')
def building_pricelist(request, object_id, type_id):
    b = Building.objects.get(pk = object_id)
    InlineFormSet = inlineformset_factory(Pricelist, ParkingCost, fields=('type','cost'), can_delete=False, extra=5, max_num=5)
    if request.method == 'POST':
        form = PricelistForm(request.POST, instance=b.pricelist)
        formset = InlineFormSet(request.POST, instance = b.pricelist)
        updateForm = PricelistUpdateForm(request.POST, prefix='upd')
        if form.is_valid():
            form.save()
            if formset.is_valid():
                formset.save()
        if updateForm.is_valid():
            action, value, date = (updateForm.cleaned_data['action'], updateForm.cleaned_data['value'],
                                   updateForm.cleaned_data['date'])
            pricelist_types = updateForm.cleaned_data['all_pricelists'] and Pricelist.objects.all() or [updateForm.cleaned_data['pricelisttype']]
            houses = [k.replace('house-','') for k in request.POST if k.startswith('house-')]
            for id in houses:
                h = House.objects.get(pk=id)
                for type in pricelist_types:
                    f = h.versions.filter(type=type)
                    if f.count() == 0: continue
                    price = f.latest().price
                    new = HouseVersion(house=h, type=type, date = date)
                    if action == PricelistUpdateForm.Add:
                        new.price = price + value
                    elif action == PricelistUpdateForm.Precentage:
                        new.price = price * (100 + value) / 100
                    elif action == PricelistUpdateForm.Multiply:
                        new.price = price * value
                    new.save()
    else:
        form = PricelistForm(instance = b.pricelist)
        formset = InlineFormSet(instance = b.pricelist)
        updateForm = PricelistUpdateForm(prefix='upd')
    
    houses = b.houses.all()
    for house in houses:
        versions = house.versions.filter(type__id = type_id)
        house.price = versions.count() > 0 and versions.latest().price or None
        
    return render(request, 'Management/building_pricelist.html',
                              {'form': form, 'formset': formset, 'updateForm':updateForm, 
                               'houses' : houses,
                               'type':PricelistType.objects.get(pk=type_id), 
                               'types':PricelistType.objects.all()}, 
                              )

@permission_required('Management.change_pricelist')
def building_pricelist_pdf(request, object_id, type_id):
    type = request.GET.get('type', '')
    b = Building.objects.get(pk = object_id)
    pricelist_type = PricelistType.objects.get(pk = type_id)
    houses = b.houses.all()
    if type == 'avaliable':
        houses = b.houses.avalible()
    for h in houses:
        try:
            h.price = h.versions.filter(type__id = type_id).latest().price
        except HouseVersion.DoesNotExist:
            h.price = None
    
    title = u'מחירון לפרוייקט %s' % str(b.project)
    subtitle = u'בניין %s - %s' % (b.num, str(pricelist_type))
    
    q = HouseVersion.objects.filter(house__building = b, type=pricelist_type)
    
    if q.count() > 0:
        subtitle += u' לתאריך ' + q.latest().date.strftime('%d/%m/%Y')

    writer = PricelistWriter(b.pricelist, houses, title, subtitle)
    
    return build_and_return_pdf(writer)

class BuildingUpdate(PermissionRequiredMixin, UpdateView):
    model = Building
    form_class = BuildingForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_building'

@permission_required('Management.building_clients')
def building_clients(request, object_id):
    building = Building.objects \
        .prefetch_related('houses__type','houses__parkings','houses__storages') \
        .get(pk = object_id)

    houses = building.houses.all()

    total_sale_price = 0
    
    for house in houses:
        sale = house.get_sale()
        if sale:
            total_sale_price += sale.price
        try:
            house.price = house.versions.company().latest().price
        except HouseVersion.DoesNotExist:
            house.price = None

    return render(request, 'Management/building_clients.html',
        { 'object':building, 'houses':houses, 'total_sale_price':total_sale_price})

@permission_required('Management.building_clients_pdf')
def building_clients_pdf(request, object_id):
    b = Building.objects.get(pk = object_id)
    houses = b.houses.sold()
    for h in houses:
        try:
            h.price = h.versions.company().latest().price
        except HouseVersion.DoesNotExist:
            h.price = None
    
    title = u'מצבת רוכשים לפרוייקט %s' % str(b.project)
    subtitle = u'בניין %s' % b.num
    
    writer = BuildingClientsWriter(houses, title, subtitle)

    return build_and_return_pdf(writer)

@permission_required('Management.add_project')
def project_add(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        ecForm = ExistContactForm(request.POST)
        contactForm = ContactForm(request.POST, prefix='contact')
        if form.is_valid():
            project = form.save()
            if contactForm.is_valid():
                project.demand_contact = contactForm.save()
            elif ecForm.is_valid():
                project.demand_contact = ecForm.cleaned_data['contact']
            project.save()
            return HttpResponseRedirect('../%s/' % project.id)
        else:
            ecForm = ExistContactForm()
    else:
        form = ProjectForm()
        ecForm = ExistContactForm()
        contactForm = ContactForm(prefix='contact')
    return render(request, 'Management/project_add.html',
                              { 'form':form,'ecForm':ecForm, 'contactForm':contactForm },
                              )

@permission_required('Management.change_salecommissiondetail')
def salecommissiondetail_edit(request, sale_id):
    sale = Sale.objects.get(pk=sale_id)
    InlineFormSet = inlineformset_factory(Sale, SaleCommissionDetail, fields=('commission','value','sale'), can_delete=False)
    if request.method == 'POST':
        formset = InlineFormSet(request.POST, instance=sale)
        if formset.is_valid():
            formset.save()
    else:
        formset = InlineFormSet(instance=sale)
        
    return render(request, 'Management/objectset_edit.html', 
                              { 'formset':formset },
                              )
    
@permission_required('Management.change_project')
def project_edit(request, id):
    project = Project.objects.select_related('demand_contact','payment_contact').get(pk=id)
    details = project.details or ProjectDetails()
    transactions = []
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        detailsForm = ProjectDetailsForm(request.POST, instance=details, prefix='det')
        if form.is_valid():
            form.save()
        if detailsForm.is_valid():
            project.details = detailsForm.save()
            project.save()
    else:
        form = ProjectForm(instance=project)
        detailsForm = ProjectDetailsForm(instance=details, prefix='det')
        commission = project.commissions.get()
        transactionsForm = TransactionUpdateHistory.objects.filter(transaction_id = commission.id, transaction_type = 1)

        for value in transactionsForm :
            field_value = value.field_value
            field_name  = value.field_name

            try :
                if field_name == 'max' :
                    field_name  = u'מקסימום עמלה'
                    field_value = field_value + ' %'
                else :
                    field_value = commaise( int( field_value ) ) + ' ש"ח'
            except :
                pass

            transactions.append( [ gettext( field_name ), field_value, str( value.timestamp ).split(' ')[0] ] )

    context = {
        'form':form, 
        'detailsForm':detailsForm, 
        'project':project, 
        'history':transactions
    }

    return render(request, 'Management/project_edit.html', context)
  
class ProjectCommissionUpdate(PermissionRequiredMixin, UpdateView):
    model = ProjectCommission
    form_class = ProjectCommissionForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_projectcommission'

class ProjectEndUpdate(PermissionRequiredMixin, UpdateView):
    model = Project
    form_class = ProjectEndForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_project'

@login_required  
def project_commission_del(request, project_id, commission):
    project_commission = ProjectCommission.objects.get(project_id=project_id)

    for field in project_commission._meta.fields:
        if abbrevate(field.name) == commission:
            obj = getattr(project_commission, field.name)
            break

    #unlink commission from project
    setattr(project_commission, commission, None)
    project.commissions.save()

    #delete commission
    obj.delete()

    return redirect('project-edit', project_id)

@login_required
def project_commission_add(request, project_id, commission):
    import inspect
    module = inspect.getmodule(project_commission_add)
    view_func = getattr(module, 'project_' + commission)
    return view_func(request, project_id)

@login_required
def employee_commission_del(request, employee_id, project_id, commission):
    employee = Employee.objects.get(pk = employee_id)
    c = employee.commissions.filter(project__id = project_id)[0]
    for field in c._meta.fields:
        if abbrevate(field.name) == commission:
            obj = getattr(c, field.name)
            #unlink commission from employee
            setattr(c, field.name, None)
            c.save()
            break

    #delete commission
    obj.delete()
    return HttpResponseRedirect('/employees/%s' % employee.id)

def abbrevate(s):
    s2 = ''
    for part in s.split('_'):
        s2 += part[0]
    return s2

@login_required
def employee_commission_add(request, employee_id, project_id, commission):
    import inspect
    module = inspect.getmodule(employee_commission_add)
    view_func = getattr(module, 'employee_' + commission)
    return view_func(request, employee_id, project_id)
        
@permission_required('Management.add_cvarprecentage')
def project_cvp(request, project_id):
    c = ProjectCommission.objects.get(project__id = project_id)
    cvp = c.c_var_precentage or CVarPrecentage()
    InlineFormSet = inlineformset_factory(CVarPrecentage, CPrecentage, fields=('precentage',), extra = 10, max_num = 10)

    if request.method == "POST":
        formset = InlineFormSet(request.POST, instance=cvp)
        form = CVarPrecentageForm(request.POST, instance=cvp)
        if formset.is_valid() and form.is_valid():
            c.c_var_precentage = form.save()
            c.save()
            formset.save()
    else:
        formset = InlineFormSet(instance=cvp)
        form = CVarPrecentageForm(instance=cvp)
    return render(request, 'Management/commission_inline.html', 
                              { 'formset':formset,'form':form, 'show_house_num':True },
                              )

@permission_required('Management.add_cvarprecentagefixed')
def project_cvpf(request, project_id):
    c = ProjectCommission.objects.get(project__id = project_id)
    cvpf = c.c_var_precentage_fixed or CVarPrecentageFixed()
    
    if request.method == 'POST':
        form = CVarPrecentageFixedForm(request.POST, instance = cvpf)
        if form.is_valid():
            c.c_var_precentage_fixed = form.save()
            c.save()
    else:
        form = CVarPrecentageFixedForm(instance= cvpf)
            
    return render(request, 'Management/object_edit.html', 
                              { 'form':form },
                              )
    
@permission_required('Management.add_czilber')
def project_cz(request, project_id):
    c = ProjectCommission.objects.get(project__id = project_id)
    cz = c.c_zilber or CZilber()
    
    if request.method == 'POST':
        form = CZilberForm(request.POST, instance= cz)
        if form.is_valid():
            c.c_zilber = form.save()
            c.save()
    else:
        form = CZilberForm(instance= cz)
            
    return render(request, 'Management/object_edit.html', 
                              { 'form':form },
                              )

@permission_required('Management.add_bdiscountsaveprecentage')
def project_bdsp(request, project_id):
    c = ProjectCommission.objects.get(project__id = project_id)
    bdsp = c.b_discount_save_precentage or BDiscountSavePrecentage()
    
    if request.method == 'POST':
        form = BDiscountSavePrecentageForm(request.POST, instance = bdsp)
        if form.is_valid():
            c.b_discount_save_precentage = form.save()
            c.save()
    else:
        form = BDiscountSavePrecentageForm(instance=bdsp)
            
    return render(request, 'Management/object_edit.html', 
                              { 'form':form },
                              )
    
@permission_required('Management.add_contact')
def project_contact(request, project_id, demand=False, payment=False):
    project = Project.objects.get(pk=project_id)
    if request.method == 'POST':
        form = ContactForm(request.POST)
        existForm = ExistContactForm(request.POST)
        if existForm.is_valid():
            c = existForm.cleaned_data['contact']
            form = ContactForm()
        elif form.is_valid():
            existForm = ExistContactForm(initial={'contact':(demand and project.demand_contact and project.demand_contact.id) or (payment and project.payment_contact and project.payment_contact.id)})
            c = form.save()
        else:
            c=None
        if c:
            if demand:
                project.demand_contact = c
            elif payment:
                project.payment_contact = c
            else:
                project.contacts.add(c)
            project.save()
    else:
        form = ContactForm()
        existForm = ExistContactForm(initial={'contact':(demand and project.demand_contact and project.demand_contact.id) or (payment and project.payment_contact and project.payment_contact.id)})
        
    return render(request, 'Management/project_contact_edit.html', 
                              { 'form':form, 'existForm':existForm },
                              )

@permission_required('Management.change_employee')
def employee_project_add(request, employee_id):
    if request.method == 'POST':
        form = EmployeeAddProjectForm(request.POST)
        if form.is_valid():
            project, employee, start_date = form.cleaned_data['project'], form.cleaned_data['employee'], form.cleaned_data['start_date']
            # check if the employee already has an open commission for this project
            open_commissions = employee.commissions.filter(project = project, end_date = None)
            if len(open_commissions) == 0:
                employee.projects.add(project)

                # create EPCommission
                commission = EPCommission(employee=employee, project=project, start_date=start_date)
                commission.save()

                # add commission to employee
                employee.commissions.add(commission)
            else:
                # TODO: something
                pass
    else:
        form = EmployeeAddProjectForm(initial={'employee':employee_id})
    return render(request, 'Management/object_edit.html', 
                              { 'form':form, 'title':u'העסקה בפרוייקט חדש' },
                              )

@permission_required('Management.change_employee')
def employee_project_remove(request, employee_id, project_id):
    project = Project.objects.get(pk = project_id)
    employee = Employee.objects.get(pk = employee_id)
    if request.method == 'POST':
        form = EmployeeRemoveProjectForm(request.POST)
        if form.is_valid():
            project, employee, end_date = form.cleaned_data['project'], form.cleaned_data['employee'], form.cleaned_data['end_date']
            for epc in employee.commissions.filter(project=project):
                epc.end_date = end_date
                epc.save()
            employee.projects.remove(project)
    else:
        form = EmployeeRemoveProjectForm(initial={'employee':employee_id, 'project':project.id})
    return render(request, 'Management/object_edit.html', 
                              { 'form':form, 'title':u'סיום העסקה בפרוייקט' },
                              )

class ContactListView(LoginRequiredMixin, ListView):
    model = Contact
    template_name = 'Management/contact_list.html'

class ContactCreate(PermissionRequiredMixin, CreateView):
    model = Contact
    form_class = ContactForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_contact'

class ContactUpdate(PermissionRequiredMixin, UpdateView):
    model = Contact
    form_class = ContactForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_contact'

@permission_required('Management.change_contact')
def project_removecontact(request, id, project_id):
    project = Project.objects.get(pk=project_id)
    contact = Contact.objects.get(pk=id)
    
    for attr in ['projects','projects_demand','projects_payment']:
        attr_value = getattr(contact, attr)
        try:
            attr_value.remove(project)
        except:
            pass

    return redirect('project-edit', project_id)

@permission_required('Management.delete_contact')
def project_deletecontact(request, id, project_id):
    contact = Contact.objects.get(pk=id)
    contact.projects.clear()
    contact.projects_demand.clear()
    contact.projects_payment.clear()
    contact.delete()
    
    return redirect('project-edit', project_id)
    
@permission_required('Management.delete_contact')
def contact_delete(request, project_id, id):
    contact = Contact.objects.get(pk=id)
    contact.projects.clear()
    contact.projects_demand.clear()
    contact.projects_payment.clear()
    contact.delete()
    return HttpResponseRedirect('/contacts')

@permission_required('Management.add_attachment')
def obj_add_attachment(request, obj_id, model):
    content_type = ContentType.objects.get_for_model(model)
    obj = content_type.get_object_for_this_type(pk = obj_id)
    if request.method == 'POST':
        form = AttachmentForm(request.POST, request.FILES, initial={'is_private':False})
        attachment = form.instance
        attachment.user_added = request.user
        attachment.file = request.FILES['file']
        attachment.content_type = content_type
        attachment.object_id = obj_id
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('../attachments')
    else:
        form = AttachmentForm()
    
    return render(request, 'Management/attachment_add.html',
                              {'form': form, 'obj':obj}, )

@login_required
def obj_attachments(request, obj_id, model):
    content_type = ContentType.objects.get_for_model(model)
    obj = content_type.get_object_for_this_type(pk = obj_id)
    attachments = Attachment.objects.filter(content_type = content_type, object_id = obj_id)

    if not request.user.is_staff:
        attachments = attachments.filter(is_private=False)
        
    return render(request, 'Management/object_attachment_list.html', 
                              {'attachments': attachments, 'obj':obj}, 
                              )

@permission_required('Management.add_reminder')
def obj_add_reminder(request, obj_id, model):
    obj = model.objects.get(pk= obj_id)
    if request.method == 'POST':
        form = ReminderForm(data= request.POST)
        if form.is_valid():
            r = form.save()
            obj.reminders.add(r)
            return HttpResponseRedirect('reminders')
    else:
        form = ReminderForm(initial={'status':ReminderStatusType.Added})
    
    return render(request, 'Management/object_edit.html',
                              {'form': form}, )

@login_required
def obj_reminders(request, obj_id, model):
    obj = model.objects.get(pk = obj_id)
    return render(request, 'Management/reminder_list.html',
                              {'reminders': obj.reminders}, )

class ReminderUpdate(PermissionRequiredMixin, UpdateView):
    model = Reminder
    form_class = ReminderForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_reminder'

@permission_required('Management.delete_reminder')
def reminder_del(request, id):
    r = Reminder.objects.get(pk= id)
    if r.statuses.latest().type.id == ReminderStatusType.Deleted:
        return HttpResponse('reminder is already deleted')
    else:
        r.delete()
        return HttpResponse('ok')
    
@permission_required('Management.change_reminder')
def reminder_do(request, id):
    r = Reminder.objects.get(pk= id)
    if r.statuses.latest().type.id == ReminderStatusType.Done:
        return HttpResponse('reminder is already done')
    else:
        r.do()
        return HttpResponse('ok')

@login_required
def project_buildings(request, project_id):
    p = Project.objects.get(pk = project_id)
    buildings = p.buildings.filter(is_deleted=False)
    total_signed_houses, total_houses, total_avalible_houses, total_sold_houses = 0,0,0,0
    for b in buildings:
        total_houses = total_houses + b.house_count
        total_signed_houses += len(b.houses.signed())
        total_sold_houses += len(b.houses.sold())
        total_avalible_houses += len(b.houses.avalible())

    return render(request, 'Management/building_list.html', 
                              { 'buildings' : buildings,'total_houses':total_houses,'project':p,
                               'total_signed_houses':total_signed_houses, 'total_avalible_houses':total_avalible_houses, 'total_sold_houses':total_sold_houses},
                              )

@permission_required('Management.add_building')
def building_add(request, project_id=None):
    if request.method == 'POST':
        form = BuildingForm(request.POST)
        if form.is_valid():
            building = form.save()
            return HttpResponseRedirect('../buildings')
    else:
        form = BuildingForm()
        if project_id:
            project = Project.objects.get(pk=project_id)
            form.initial = {'project':project.id}
            
    return render(request, 'Management/object_edit.html', 
                              {'form' : form})

@permission_required('Management.add_parking')
def building_addparking(request, building_id = None):
    if request.method == 'POST':
        form = ParkingForm(request.POST)
        if form.is_valid():
            form.save()
    else:
        form = ParkingForm()
        if building_id:
            b = Building.objects.get(pk=building_id)
            form.initial = {'building':b.id}
            form.fields['house'].queryset = b.houses.all()
    return render(request, 'Management/object_edit.html', 
                              {'form' : form},
                              )

class ParkingUpdate(PermissionRequiredMixin, UpdateView):
    model = Parking
    form_class = ParkingForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_parking'

@permission_required('Management.add_storage')
def building_addstorage(request, building_id = None):
    if request.method == 'POST':
        form = StorageForm(request.POST)
        if form.is_valid():
            form.save()
    else:
        form = StorageForm()
        if building_id:
            b = Building.objects.get(pk=building_id)
            form.initial = {'building':b.id}
            form.fields['house'].queryset = b.houses.all()
    return render(request, 'Management/object_edit.html', {'form' : form}, )
        
class StorageUpdate(PermissionRequiredMixin, UpdateView):
    model = Storage
    form_class = StorageForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_storage'

@permission_required('Management.delete_building')
def building_delete(request, building_id):
    building = Building.objects.get(pk = building_id)
    building.is_deleted = True
    building.save()
    return HttpResponse('ok')
        
@permission_required('Management.copy_building')
def building_copy(request, building_id):
    building = Building.objects.get(pk = building_id)
    if request.method == 'POST':
        form = CopyBuildingForm(request.POST)
        if form.is_valid():
            include_houses, include_house_prices, include_parkings, include_storages = (form.cleaned_data['include_houses'],
                                                                                        form.cleaned_data['include_house_prices'],
                                                                                        form.cleaned_data['include_parkings'],
                                                                                        form.cleaned_data['include_storages'])
            
            if building.pricelist:
                new_pricelist = common.clone(building.pricelist, False)
                new_pricelist.save()
            else:
                new_pricelist = None
            
            new_building = common.clone(building, False)
            new_building.num = form.cleaned_data['new_building_num']
            new_building.pricelist = new_pricelist
            new_building.save()
            
            if include_parkings:
                for parking in building.parkings.filter(house__isnull=True):
                    new_parking = common.clone(parking, False)
                    new_parking.building = new_building
                    new_parking.save()
                    
            if include_storages:
                for storage in building.storages.filter(house__isnull=True):
                    new_storage = common.clone(storage, False)
                    new_storage.building = new_building
                    new_storage.save()
                                
            if include_houses:
                for house in building.houses.all():
                    new_house = common.clone(house, False)
                    new_house.building = new_building
                    new_house.save()
                    
                    if include_house_prices:
                        for house_version in house.versions.all():
                            new_house_version = common.clone(house_version, False)
                            new_house_version.house = new_house
                            new_house_version.save()
                            
                    if include_parkings:
                        for parking in house.parkings.all():
                            new_parking = common.clone(parking, False)
                            new_parking.house = new_house
                            new_parking.building = new_building
                            new_parking.save()
                            
                    if include_storages:
                        for storage in house.storages.all():
                            new_storage = common.clone(storage, False)
                            new_storage.house = new_house
                            new_storage.building = new_building
                            new_storage.save()
                    
            return HttpResponseRedirect(reverse(project_buildings, args=[building.project_id]))
    else:
        form = CopyBuildingForm()
        form.fields['building'].queryset = building.project.buildings.filter(is_deleted=False)
        form.fields['building'].initial = building.id
        
    return render(request, 'Management/object_edit.html',
                              {'form' : form, 'title':gettext('copy_building')}, 
                              )
    
@permission_required('Management.add_house')
def building_addhouse(request, type_id, building_id):
    b = Building.objects.get(pk=building_id)
    if request.method == 'POST':
        form = HouseForm(type_id, data = request.POST)
        form.instance.building = b
        if form.is_valid():
            form.save(type_id)
            if 'addanother' in request.POST:
                next_url = reverse(building_addhouse, args=[building_id, type_id])
            elif 'finish' in request.POST:
                next_url = reverse(building_pricelist, args=[building_id, type_id])
            else:
                next_url = reverse(house_edit, args=[form.instance.id, type_id])
            return HttpResponseRedirect(next_url)
    else:
        form = HouseForm(type_id)
        form.instance.building = b
        if b.is_cottage():
            form.initial['type'] = HouseType.Cottage

    ps = Parking.objects.filter(building = b)
    ss = Storage.objects.filter(building = b)
    for f in ['parking1','parking2','parking3']:
        form.fields[f].queryset = ps
    for f in ['storage1','storage2']:
        form.fields[f].queryset = ss
    return render(request, 'Management/house_edit.html', 
                              {'form' : form, 'type':PricelistType.objects.get(pk = type_id) },
                              )

@permission_required('Management.house_versionlog')
def house_version_log(request, building_id, id ,type_id):
    house = House.objects.get(pk=id)
    pricelist_type = PricelistType.objects.get(pk=type_id)
    query = HouseVersion.objects.filter(house__id = id, type__id = type_id)
    versions = list(query)
    previous_version = None
    for version in versions:
        if previous_version:
            version.diff_amount = version.price - previous_version.price
            version.diff_precentage = float(version.price) / previous_version.price * 100 - 100
        previous_version = version
    
    return render(request, 'Management/house_version_log.html', 
                              {'title':u'מחירי %s עבור דירה %s' % (pricelist_type, house.num),
                               'versions':versions })
    
@permission_required('Management.change_house')
def house_edit(request, building_id, id ,type_id):
    h = House.objects.get(pk=id)
    b = h.building
    if request.method == 'POST':
        form = HouseForm(type_id, data = request.POST, instance = h)
        if form.is_valid():
            form.save(type_id)
            if 'addanother' in request.POST:
                return HttpResponseRedirect('../../addhouse/type%s' % type_id)
            elif 'finish' in request.POST:
                return HttpResponseRedirect('../../pricelist/type%s' % type_id)
    else:
        form = HouseForm(type_id, instance = h)
        if b.is_cottage():
            form.initial['type'] = HouseType.Cottage

    ps = Parking.objects.filter(building = b)
    ss = Storage.objects.filter(building = b)
    for f in ['parking1','parking2','parking3']:
        form.fields[f].queryset = ps
    for f in ['storage1','storage2']:
        form.fields[f].queryset = ss
    return render(request, 'Management/house_edit.html', 
                              {'form' : form, 'type':PricelistType.objects.get(pk = type_id) },
                              )

@permission_required('Management.delete_house')
def house_delete(request, building_id, type_id, house_id):
    house = House.objects.get(pk = house_id)
    house.delete()
    return HttpResponseRedirect("../../../type%s" % type_id)
        
@permission_required('Management.add_loan')
def employee_addloan(request, employee_id):
    employee = Employee.objects.get(pk = employee_id)
    if request.method=='POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            form.save() 
    else:
        form = LoanForm(initial={'employee':employee.id})
    
    return render(request, 'Management/object_edit.html',
                              {'form' : form}, )
    
@permission_required('Management.add_loanpay')
def employee_loanpay(request, employee_id):
    e = Employee.objects.get(pk=employee_id)
    if request.method == 'POST':
        form = LoanPayForm(request.POST)
        if form.is_valid():
            form.instance.employee = e
            form.save()
    else:
        form = LoanPayForm(initial={'employee':e.id})
    return render(request, 'Management/object_edit.html',
                              {'form' : form}, )

class EmployeeCreate(PermissionRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    permission_required = 'Management.change_employee'

class EmployeeUpdate(PermissionRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    permission_required = 'Management.change_employee'

@permission_required('Management.change_employee')
def employee_end(request, object_id):
    employee = EmployeeBase.objects.get(pk=object_id)
    if request.method == 'POST':
        form = EmployeeEndForm(request.POST, instance = employee)
        if form.is_valid():
            form.save()

            employee_derived = employee.derived

            if isinstance(employee_derived, Employee):
                for epcommission in employee_derived.commissions.all():
                    if not epcommission.end_date:
                        epcommission.end_date = employee.work_end
                        epcommission.save()
            elif isinstance(employee_derived, NHEmployee):
                for nhbranchemployee in employee_derived.nhbranchemployee_set.all():
                    if not nhbranchemployee.end_date:
                        nhbranchemployee.end_date = employee.work_end
                        nhbranchemployee.save()
    else:
        form = EmployeeEndForm(instance = employee)
        
    return render(request, 'Management/object_edit.html',
                              {'form' : form}, )

@permission_required('Management.add_loan')
def nhemployee_addloan(request, employee_id):
    employee = NHEmployee.objects.get(pk = employee_id)
    if request.method=='POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            form.save() 
    else:
        form = LoanForm(initial={'employee':employee.id})
    
    return render(request, 'Management/object_edit.html',
                              {'form' : form}, )
    
@permission_required('Management.add_loanpay')
def nhemployee_loanpay(request, employee_id):
    e = NHEmployee.objects.get(pk=employee_id)
    if request.method == 'POST':
        form = LoanPayForm(request.POST)
        if form.is_valid():
            form.instance.nhemployee = e
            form.save()
    else:
        form = LoanPayForm(initial={'employee':e.id})
    return render(request, 'Management/object_edit.html',
                              {'form' : form}, )

@permission_required('Management.add_employmentterms')
def employee_employmentterms(request, id, model):
    employee = model.objects.get(pk = id)
    terms = employee.employment_terms or EmploymentTerms()
    terms_dict = terms.__dict__.copy()

    allow_history = [ 'salary_base', 'safety', 'tax_deduction_source_precentage' ]

    if request.method == "POST":
        form = EmploymentTermsForm(request.POST, instance=terms)
        if form.is_valid():
            form_dict = form.instance.__dict__

            for item in terms._meta.fields :
                name = item.name

                if name in form_dict and name in allow_history :
                    value     = form_dict.get(name)
                    value_len = value != None and len(str(value)) or 0

                    if value_len > 0 and ( terms_dict.get(name) != value ):
                        TransactionUpdateHistory( transaction_type='0', transaction_id=id, field_name=name, field_value=(value and value or ''), timestamp=datetime.now() ).save()

            employee.employment_terms = form.save()
            employee.save()
    else:
        form = EmploymentTermsForm(instance=terms)
    return render(request, 'Management/object_edit.html', 
                              { 'form':form },
                              )
    
@permission_required('Management.add_cvar')
def employee_cv(request, employee_id, project_id):
    pc = EPCommission.objects.get(employee__id = employee_id, project__id = project_id, end_date__isnull = True)
    cv = pc.c_var or CVar()
    InlineFormSet = inlineformset_factory(CVar, CAmount, fields=('amount',), extra = 10, max_num = 10)
    
    if request.method == "POST":
        formset = InlineFormSet(request.POST, instance=cv)
        form = CVarForm(request.POST, instance=cv)
        if formset.is_valid() and form.is_valid():
            pc.c_var = form.save()
            pc.save()
            formset.save()
    else:
        formset = InlineFormSet(instance=cv)
        form = CVarForm(instance=cv)
    return render(request, 'Management/commission_inline.html', 
                              { 'formset':formset,'form':form, 'show_house_num':True },
                              )
    
@permission_required('Management.add_cvarprecentage')
def employee_cvp(request, employee_id, project_id):
    pc = EPCommission.objects.get(employee__id = employee_id, project__id = project_id, end_date__isnull = True)
    cvp = pc.c_var_precentage or CVarPrecentage()
    InlineFormSet = inlineformset_factory(CVarPrecentage, CPrecentage, fields=('precentage',), extra = 10, max_num = 10)
    
    if request.method == "POST":
        formset = InlineFormSet(request.POST, instance=cvp)
        form = CVarPrecentageForm(request.POST, instance=cvp)
        if formset.is_valid() and form.is_valid():
            pc.c_var_precentage = form.save()
            pc.save()
            formset.save()
    else:
        formset = InlineFormSet(instance=cvp)
        form = CVarPrecentageForm(instance=cvp)
    return render(request, 'Management/commission_inline.html', 
                              { 'formset':formset,'form':form, 'show_house_num':True },
                              )
    
@permission_required('Management.add_cbyprice')
def employee_cbp(request, employee_id, project_id):
    pc = EPCommission.objects.get(employee__id = employee_id, project__id = project_id, end_date__isnull = True)
    cbp = pc.c_by_price or CByPrice()
    InlineFormSet = inlineformset_factory(CByPrice, CPriceAmount,CPriceAmountForm, extra = 10, max_num = 10)
    if request.method == "POST":
        formset = InlineFormSet(request.POST, instance=cbp)
        form = CByPriceForm(request.POST, instance=cbp)
        if formset.is_valid() and form.is_valid():
            pc.c_by_price = form.save()
            pc.save()
            formset.save()
    else:
        formset = InlineFormSet(instance = cbp)
        form = CByPriceForm(instance = cbp)
        
    return render(request, 'Management/commission_inline.html', 
                              { 'form':form, 'formset':formset, 'show_house_num':False },
                              )

@permission_required('Management.add_bsalerate')
def employee_bsr(request, employee_id, project_id):
    pc = EPCommission.objects.get(employee__id = employee_id, project__id = project_id, end_date__isnull = True)
    bsr = pc.b_sale_rate or BSaleRate()
    InlineFormSet = inlineformset_factory(BSaleRate, SaleRateBonus, fields=('house_count','amount'), extra = 10, max_num = 10)
    if request.method == "POST":
        formset = InlineFormSet(request.POST, instance=bsr)
        form = BSaleRateForm(request.POST, instance=bsr)
        if formset.is_valid() and form.is_valid():
            pc.b_sale_rate = form.save()
            pc.save()
            formset.save()
    else:
        formset = InlineFormSet(instance = bsr)
        form = BSaleRateForm(instance = bsr)
        
    return render(request, 'Management/commission_inline.html', 
                              { 'form':form, 'formset':formset, 'show_house_num':False},
                              )
    
@permission_required('Management.add_bhousetype')
def employee_bht(request, employee_id, project_id):
    pc = EPCommission.objects.get(employee__id = employee_id, project__id = project_id, end_date__isnull = True)
    htb = pc.b_house_type or BHouseType()
    InlineFormSet = inlineformset_factory(BHouseType, HouseTypeBonus, fields=('type','amount'), extra = 10, max_num = 10)
    if request.method == "POST":
        formset = InlineFormSet(request.POST, instance=htb)
        form = BHouseTypeForm(request.POST, instance=htb)
        if formset.is_valid() and form.is_valid():
            pc.b_house_type = form.save()
            pc.save()
            formset.save()
    else:
        formset = InlineFormSet(instance=htb)
        form = BHouseTypeForm(instance = htb)
        
    return render(request, 'Management/commission_inline.html', 
                              { 'form':form, 'formset':formset, 'show_house_num':False},
                              )
    
@permission_required('Management.add_bdiscountsave')
def employee_bds(request, employee_id, project_id):
    pc = EPCommission.objects.get(employee__id = employee_id, project__id = project_id, end_date__isnull = True)
    bds = pc.b_discount_save or BDiscountSave()
    
    if request.method == 'POST':
        form = BDiscountSaveForm(request.POST, instance = bds)
        if form.is_valid():
            pc.b_discount_save = form.save()
            pc.save()
    else:
        form = BDiscountSaveForm(instance=bds)
            
    return render(request, 'Management/object_edit.html', 
                              { 'form':form },
                              )

@permission_required('Management.add_bdiscountsaveprecentage')
def employee_bdsp(request, employee_id, project_id):
    pc = EPCommission.objects.get(employee__id = employee_id, project__id = project_id, end_date__isnull = True)
    bdsp = pc.b_discount_save_precentage or BDiscountSavePrecentage()
    
    if request.method == 'POST':
        form = BDiscountSavePrecentageForm(request.POST, instance = bdsp)
        if form.is_valid():
            pc.b_discount_save_precentage = form.save()
            pc.save()
    else:
        form = BDiscountSavePrecentageForm(instance=bdsp)
            
    return render(request, 'Management/object_edit.html', 
                              { 'form':form },
                              )

class NHBranchEmployeeCreate(PermissionRequiredMixin, CreateView):
    model = NHBranchEmployee
    form_class = NHBranchEmployeeForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.add_nhbranchemployee'

    def get_initial(self):
        initial = super().get_initial()
        branch_id = self.kwargs.get('nhbranch_id')
        initial['nhbranch'] = branch_id
        return initial

class NHBranchEmployeeUpdate(PermissionRequiredMixin, UpdateView):
    model = NHBranchEmployee
    form_class = NHBranchEmployeeForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_nhbranchemployee'

@login_required
def json_buildings(request, project_id):
    data = serializers.serialize('json', Project.objects.get(pk= project_id).non_deleted_buildings(), 
                                 fields=('id','name','num'))
    return HttpResponse(data)

@login_required
def json_employees(request, project_id):
    l = [EmployeeBase.objects.get(pk=e.id) for e in Project.objects.get(pk=project_id).employees.active()]
    data = serializers.serialize('json', l, 
                                 fields=('id','pid','first_name','last_name'))
    return HttpResponse(data)

@login_required
def json_houses(request, building_id):
    houses = House.objects.filter(building__id = building_id)
    data = serializers.serialize('json', houses, fields=('id','num'))
    return HttpResponse(data)

@login_required
def json_house(request, house_id):
    data = serializers.serialize('json', [House.objects.get(pk= house_id),])
    return HttpResponse(data)
    house = House.objects.get(pk= house_id)
    fields = {}
    for field in ['id', 'num', 'floor', 'rooms', 'allowed_discount', 'parking', 'warehouse', 'remarks']:
        fields[field] = 1 and getattr(house, field) or ''
    fields['price'] = house.prices.latest().price
    a = '[' + str({'pk':house.id, 'model':'Management.house', 'fields':fields}) + ']'
    return HttpResponse(a)
      
class AttachmentUpdate(PermissionRequiredMixin, UpdateView):
    model = Attachment
    form_class = AttachmentForm
    template_name = 'Management/object_edit.html'
    permission_required = 'Management.change_attachment'

class AttachmentDelete(PermissionRequiredMixin, DeleteView):
    model = Attachment
    success_url = '/attachments'
    template_name = 'Management/object_confirm_delete.html'
    permission_required = 'Management.delete_attachment'

@permission_required('Management.list_attachment')
def attachment_list(request):
    attachments = Attachment.objects.all()
    
    if len(request.GET):
        project_select_form = ProjectSelectForm(request.GET, prefix = 'project')
        employee_select_form = EmployeeSelectForm(request.GET, prefix = 'employee')
        demand_select_form = DemandSelectForm(request.GET, prefix = 'demand')
        
        object_id, content_type = None, None
        
        if 'project' in request.GET:
            if project_select_form.is_valid():
                project = project_select_form.cleaned_data['project']
                model = project.__class__
                content_type = ContentType.objects.get_for_model(model)
                object_id = project.id
        elif 'employee' in request.GET:
            if employee_select_form.is_valid():
                employee = employee_select_form.cleaned_data['employee']
                model = employee.__class__
                content_type = ContentType.objects.get_for_model(model)
                object_id = employee.id
        elif 'demand' in request.GET:
            if demand_select_form.is_valid():
                kwargs = demand_select_form.cleaned_data
                demand = Demand.objects.get(**kwargs)
                model = demand.__class__
                content_type = ContentType.objects.get_for_model(model)
                object_id = demand.id
        
        attachments = attachments.filter(content_type = content_type, object_id = object_id)
    else:
        project_select_form = ProjectSelectForm(prefix = 'project')
        employee_select_form = EmployeeSelectForm(prefix = 'employee')
        demand_select_form = DemandSelectForm(prefix = 'demand')
          
    return render(request, 'Management/attachment_list.html', 
                              {'project_select_form':project_select_form, 'employee_select_form':employee_select_form,
                               'demand_select_form':demand_select_form, 'attachments':attachments },
                              )

@permission_required('Management.add_attachment')
def attachment_add(request):
    if request.method == "POST":
        form = AttachmentForm(request.POST, request.FILES)
        project_select_form = ProjectSelectForm(request.POST, prefix = 'project')
        employee_select_form = EmployeeSelectForm(request.POST, prefix = 'employee')
        demand_select_form = DemandSelectForm(request.POST, prefix = 'demand')
        
        attachment = form.instance
        attachment.user_added = request.user
        attachment.file = request.FILES['file']
        
        if 'project' in request.POST:
            if project_select_form.is_valid():
                attachment.content_type = ContentType.objects.get_for_model(Project)
                attachment.object_id = project_select_form.cleaned_data['project'].id
        elif 'employee' in request.POST:
            if employee_select_form.is_valid():
                attachment.content_type = ContentType.objects.get_for_model(EmployeeBase)
                attachment.object_id = employee_select_form.cleaned_data['employee'].id
        elif 'demand' in request.POST:
            if demand_select_form.is_valid():
                attachment.content_type = ContentType.objects.get_for_model(Demand)
                kwargs = demand_select_form.cleaned_data
                demand = Demand.objects.get(**kwargs)
                attachment.object_id = demand.id
        
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('/attachments')
    else:
        form = AttachmentForm()
        project_select_form = ProjectSelectForm(prefix = 'project')
        employee_select_form = EmployeeSelectForm(prefix = 'employee')
        demand_select_form = DemandSelectForm(prefix = 'demand')
        
    return render(request, 'Management/attachment_add.html', 
                              {'form':form, 'project_select_form':project_select_form, 'employee_select_form':employee_select_form,
                               'demand_select_form':demand_select_form },
                              )

@permission_required('Management.change_sale')
def sale_edit(request, id):
    sale = Sale.objects.get(pk=id)
    demand = sale.demand

    if request.POST:
        form = SaleForm(request.POST, instance = sale)
        #handles the case when the building changes, and the house is not in the queryset of the house field
        form.fields['house'].queryset = House.objects.all()
        form.fields['building'].queryset = Building.objects.all()
        if form.is_valid():
            project = form.cleaned_data['project']
            next = None

            #temp fix. should remove
            if demand.statuses.count() == 0:
                demand.feed()
            if demand.was_sent:
                #check for mods:
                if form.changed_data.count('price'):
                    try:
                        spm = sale.salepricemod
                    except SalePriceMod.DoesNotExist:
                        spm = SalePriceMod(sale = sale, old_price = sale.price) 
                    spm.save()
                    next = '/salepricemod/%s' % spm.id
                if form.changed_data.count('house'):
                    try:
                        shm = sale.salehousemod
                    except SaleHouseMod.DoesNotExist:
                        shm = SaleHouseMod(sale = sale, old_house = sale.house)
                    shm.save()
                    next = '/salehousemod/%s' % shm.id
            
            form.save()

            year, month = demand.year, demand.month

            # enrich demand
            set_demand_sale_fields([demand], year, month, year, month)
            set_demand_diff_fields([demand])

            demand.calc_sales_commission()

            employees = demand.project.employees.exclude(
                work_end__isnull=False, 
                work_end__lt=date(year, month, 1))
            
            salaries_to_calc = EmployeeSalary.objects \
                .nondeleted() \
                .filter(employee__in = employees, year = year, month = month)

            for salary in salaries_to_calc:
                salary.calculate()
                salary.save()

            if 'addanother' in request.POST:
                return HttpResponseRedirect(next or '/demands/%s/sale/add' % sale.demand.id)
            elif 'todemand' in request.POST:
                return HttpResponseRedirect(next or '/demands/%s' % sale.demand.id)
    else:
        form = SaleForm(instance= sale)
    return render(request, 'Management/sale_edit.html', 
                              {'form':form, 'year':sale.actual_demand.year, 'month':sale.actual_demand.month},
                              )    

@permission_required('Management.add_sale')
def sale_add(request, demand_id=None):
    if demand_id:
        demand = Demand.objects.get(pk = demand_id)
        year, month = demand.year, demand.month
    else:
        year, month = common.current_month().year, common.current_month().month
    if request.POST:
        form = SaleForm(request.POST)
        if form.is_valid():
            if not demand_id:
                demand, new = Demand.objects.get_or_create(year = year, month = month, project = form.cleaned_data['project'])
            form.instance.demand = demand
            form.save()
            next = None
            if demand.statuses.count() == 0:
                demand.feed()
            if demand.was_sent:
                sp = SalePre(sale = form.instance,
                             to_month = month, to_year = year,
							 employee_pay_year = month == 12 and year + 1 or year,
							 employee_pay_month = month == 12 and 1 or month)
                sp.save()
                next = '/salepre/%s' % sp.id 
            
            # enrich demand
            set_demand_sale_fields([demand], year, month, year, month)
            set_demand_diff_fields([demand])

            demand.calc_sales_commission()
            
            employees = demand.project.employees.exclude(work_end__isnull = False, work_end__lt = date(year, month, 1))
            
            salaries_to_calc = EmployeeSalary.objects \
                .nondeleted() \
                .filter(employee__in = employees, year = year, month = month)

            for salary in salaries_to_calc:
                salary.calculate()
                salary.save()
                            
            if 'addanother' in request.POST:
                return HttpResponseRedirect(next or reverse(sale_add, args=[demand_id]))
            elif 'todemand' in request.POST:
                return HttpResponseRedirect(next or '/demands/%s' % demand.id)
    else:
        form = SaleForm()
        if demand_id:
            p = demand.project
            form.fields['project'].initial = p.id
            form.fields['employee'].queryset = p.employees.all()
            form.fields['building'].queryset = p.buildings.all()
            form.fields['commission_madad_bi'].initial = demand.get_madad()
    return render(request, 'Management/sale_edit.html', 
                              {'form':form, 'year':year, 'month':month},
                              )

def demand_sale_list(request):
    demand_id = int(request.GET.get('demand_id', 0))
    project_id = int(request.GET.get('project_id', 0))
    from_year = int(request.GET.get('from_year', 0))
    from_month = int(request.GET.get('from_month', 0))
    to_year = int(request.GET.get('to_year', 0))
    to_month = int(request.GET.get('to_month', 0))

    if demand_id:
        demand = Demand.objects.get(pk=demand_id)
        project, year, month = demand.project, demand.year, demand.month

        demands = [demand]

        set_demand_sale_fields(demands, year, month, year, month)

        title = u'ריכוז מכירות לפרוייקט %s לחודש %s/%s' % (str(project), month, year)
    elif project_id:
        project = Project.objects.get(pk=project_id)
        
        demands = Demand.objects \
            .range(from_year, from_month, to_year, to_month) \
            .filter(project = project)

        set_demand_sale_fields(demands, from_year, from_month, to_year, to_month)

        title = u'ריכוז מכירות לפרוייקט %s מחודש %s/%s עד חודש %s/%s' % (str(project), from_month, from_year,
                                                                         to_month, to_year)
    else:
        raise ValueError

    sales = []
    sales_amount = 0

    for demand in demands:
        sales.extend(demand.sales_list)
        sales_amount += demand.sales_amount

    return render(request, 'Management/sale_list.html', 
        {'sales':sales, 'sales_amount':sales_amount,'title':title})

@login_required
def project_demands(request, project_id, demand_type):
    project = Project.objects.get(pk=project_id)

    all_demands = Demand.objects \
        .prefetch_related('reminders__statuses','invoices__offset','payments') \
        .select_related('project') \
        .filter(project_id=project_id) \
        .order_by('year','month')

    set_demand_diff_fields(all_demands)
    set_demand_invoice_payment_fields(all_demands)

    # exclude fully-paid demands
    demands = list(filter(lambda demand: demand.is_fully_paid == False, all_demands))

    from_year, from_month = demands[0].year, demands[0].month
    to_year, to_month = demands[-1].year, demands[-1].month

    set_demand_sale_fields(demands, from_year, from_month, to_year, to_month)
    set_demand_is_fixed(demands)
    set_demand_open_reminders(demands)

    if demand_type == 'mis-paid':
        func = lambda demand: demand.invoices_amount != None and demand.payments_amount != None and demand.diff_invoice_payment != 0
        template_name = 'Management/project_demands_mispaid.html'
    elif demand_type == 'un-paid':
        func = lambda demand: demand.invoices_amount == None and demand.payments_amount == None
        template_name = 'Management/project_demands_unpaid.html'
    elif demand_type == 'no-invoice':
        func = lambda demand: demand.invoices_amount == None and demand.payments_amount != None
        template_name = 'Management/project_demands_noinvoice.html'
    elif demand_type == 'no-payment':
        func = lambda demand: demand.invoices_amount != None and demand.payments_amount == None
        template_name = 'Management/project_demands_nopayment.html'
    elif demand_type == 'not-yet-paid':
        func = lambda demand: demand.not_yet_paid == 1
        template_name = 'Management/project_demands_noinvoice.html'

    # apply the correct filter
    demands = filter(func, demands)

    return render(request, template_name, {'demands':demands, 'project':project})

@login_required
def demand_sales(request, project_id, year, month):
    try:
        demand = Demand.objects.get(project__id = project_id, year = year, month = month)

        set_demand_sale_fields([demand], year, month, year, month)

        sales = demand.sales_list
    except Demand.DoesNotExist:
        sales = []
        
    return render(request, 'Management/sale_table.html', {'sales':sales})

@permission_required('Management.report_employee_sales')
def report_employee_sales(request):
    if len(request.GET):
        form = ProjectSeasonForm(request.GET)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            project, from_month, from_year, to_month, to_year = cleaned_data['project'], cleaned_data['from_month'], \
                cleaned_data['from_year'], cleaned_data['to_month'], cleaned_data['to_year']
                
            demands = Demand.objects \
                .range(from_year, from_month, to_year, to_month) \
                .filter(project = project)
            
            # enrich demands
            set_demand_sale_fields(demands, from_year, from_month, to_year, to_month)

            writer = EmployeeSalesWriter(project, from_month, from_year, to_month, to_year, demands)
            
            return build_and_return_pdf(writer)

    error = u'לא ניתן להפיק את הדו"ח. אנא ודא שכל הנתונים הוזנו כראוי.'
    return render(request, 'Management/error.html', {'error': error}, )
        
    

@permission_required('Management.demand_pdf')
def report_project_month(request, project_id = 0, year = 0, month = 0, demand = None):
    if not (demand or (project_id and year and month)):
        raise ValueError % 'must supply either demand or project_id, year and month'
    
    if not demand:
        demand = Demand.objects.get(project__id = project_id, year = year, month = month)
    
    set_demand_sale_fields([demand], demand.year, demand.month, demand.year, demand.month)

    if demand.sales_count == 0:
        return render(request, 'Management/error.html', {'error':u'לדרישה שנבחרה אין מכירות'}, )
    
    set_demand_diff_fields([demand])

    writer = MonthDemandWriter(demand, to_mail=False)
    
    return build_and_return_pdf(writer)

@permission_required('Management.report_projects_month')
def report_projects_month(request, year, month):

    demands = Demand.objects \
        .prefetch_related('invoices','payments') \
        .select_related('project') \
        .filter(year=year, month=month)

    set_demand_sale_fields(demands, year, month, year, month)
    set_demand_diff_fields(demands)

    title = u'ריכוז דרישות לפרוייקטים לחודש %s\%s' % (year, month)

    writer = MultipleDemandWriter(demands, title, show_month=False, show_project=True)

    return build_and_return_pdf(writer)

@permission_required('demand_season_pdf')
def report_project_season(request, project_id=None, 
    from_year=common.current_month().year, from_month=common.current_month().month, 
    to_year=common.current_month().year, to_month=common.current_month().month):

    from_date = date(int(from_year), int(from_month), 1)
    to_date = date(int(to_year), int(to_month), 1)
    
    demands = Demand.objects \
        .prefetch_related('invoices','payments') \
        .select_related('project') \
        .range(from_year, from_month, to_year, to_month) \
        .filter(project__id = project_id)

    set_demand_sale_fields(demands, from_year, from_month, to_year, to_month)
    set_demand_diff_fields(demands)

    project = Project.objects.get(pk=project_id)

    title = u'ריכוז דרישות תקופתי לפרוייקט %s' % project

    writer = MultipleDemandWriter(demands, title, show_month=True, show_project=False)

    return build_and_return_pdf(writer)

@permission_required('demand_followup_pdf')
def report_project_followup(request, project_id=None, 
    from_year=common.current_month().year, from_month=common.current_month().month, 
    to_year=common.current_month().year, to_month=common.current_month().month):
    
    project = Project.objects.get(pk = project_id)
    
    demands = Demand.objects \
        .prefetch_related('invoices','payments') \
        .select_related('project') \
        .range(from_year, from_month, to_year, to_month) \
        .filter(project__id = project_id)

    set_demand_sale_fields(demands, from_year, from_month, to_year, to_month)
    set_demand_diff_fields(demands)
    set_demand_invoice_payment_fields(demands)

    writer = DemandFollowupWriter(project, from_month, from_year, to_month, to_year, demands)
    
    return build_and_return_pdf(writer)

@permission_required('Management.demand_season')
def demand_season_list(request):
    ds = Demand.objects.none()
    total_sales_count,total_sales_amount, total_amount = 0,0,0
    from_date, to_date, project = None, None, None
    
    if len(request.GET):
        form = ProjectSeasonForm(request.GET)
        if form.is_valid():
            project = form.cleaned_data['project']

            from_year, from_month = form.cleaned_data['from_year'], form.cleaned_data['from_month']
            to_year, to_month = form.cleaned_data['to_year'], form.cleaned_data['to_month']

            from_date = date(from_year, from_month, 1)
            to_date = date(to_year, to_month, 1)
            
            ds = Demand.objects \
                .range(from_year, from_month, to_year, to_month) \
                .filter(project = project) \
                .prefetch_related('reminders__statuses') \
                .select_related('project')

            set_demand_sale_fields(ds, from_year, from_month, to_year, to_month)
            set_demand_diff_fields(ds)
            set_demand_last_status(ds)
            set_demand_is_fixed(ds)
            set_demand_open_reminders(ds)

            for d in ds:
                total_sales_count += d.sales_count
                total_sales_amount += d.sales_amount
                total_amount += d.total_amount
    else:
        form = ProjectSeasonForm()
        
    context = { 
        'demands':ds, 
        'start':from_date, 'end':to_date,
        'project':project, 'filterForm':form,
        'total_sales_count':total_sales_count,
        'total_sales_amount':total_sales_amount,
        'total_amount':total_amount
    }

    return render(request, 'Management/demand_season_list.html', context)

def demand_season_list_export(request):

    if not len(request.GET):
        return HttpResponseBadRequest()

    form = ProjectSeasonForm(request.GET)

    if not form.is_valid():
        return HttpResponseBadRequest()

    project = form.cleaned_data['project']

    from_year, from_month = form.cleaned_data['from_year'], form.cleaned_data['from_month']
    to_year, to_month = form.cleaned_data['to_year'], form.cleaned_data['to_month']

    from_date = date(from_year, from_month, 1)
    to_date = date(to_year, to_month, 1)
    
    ds = Demand.objects \
        .range(from_year, from_month, to_year, to_month) \
        .filter(project = project) \
        .prefetch_related('reminders__statuses') \
        .select_related('project')

    set_demand_sale_fields(ds, from_year, from_month, to_year, to_month)
    set_demand_diff_fields(ds)
    # set_demand_last_status(ds)
    # set_demand_is_fixed(ds)

    columns = [
        ExcelColumn("מס'"), 
        ExcelColumn('חודש'), 
        ExcelColumn("מס' מכירות", showSum=True), 
        ExcelColumn('סה"כ מכירות כולל מע"מ', 'currency', showSum=True, width=20),
        ExcelColumn('עמלה מחושב בגין שיווק', 'currency', width=20),
        ExcelColumn('תוספת קבועה', 'currency'),
        ExcelColumn('תוספת משתנה', 'currency'),
        ExcelColumn('בונוס', 'currency'),
        ExcelColumn('קיזוז', 'currency'),
        ExcelColumn('סה"כ תשלום לחברה', 'currency', showSum=True, width=20),
    ]

    rows = []

    # Iterate through all demands
    for demand in ds:
        # Define the data for each cell in the row 
        row = [
            demand.id,
            '{month}/{year}'.format(month=demand.month, year=demand.year),
            demand.sale_count,
            demand.sales_amount,
            demand.sales_commission,
            getattr(demand, 'fixed_diff', None),
            getattr(demand, 'var_diff', None),
            getattr(demand, 'bonus_diff', None),
            getattr(demand, 'fee_diff', None),
            demand.total_amount
        ]
        
        rows.append(row)

    title = 'ריכוז דרישות - {project_name}'.format(project_name=project.name)

    return ExcelGenerator().generate(title, columns, rows)

def demand_followup_export(request):
    if not len(request.GET):
        return HttpResponseBadRequest()

    form = ProjectSeasonForm(request.GET)

    if not form.is_valid():
        return HttpResponseBadRequest()

    project = form.cleaned_data['project']

    from_year, from_month = form.cleaned_data['from_year'], form.cleaned_data['from_month']
    to_year, to_month = form.cleaned_data['to_year'], form.cleaned_data['to_month']

    from_date = date(from_year, from_month, 1)
    to_date = date(to_year, to_month, 1)
    
    demands = Demand.objects \
        .prefetch_related('invoices','payments') \
        .select_related('project') \
        .range(from_year, from_month, to_year, to_month) \
        .filter(project__id = project.id)

    set_demand_sale_fields(demands, from_year, from_month, to_year, to_month)
    set_demand_diff_fields(demands)
    set_demand_invoice_payment_fields(demands)

    columns = [
        ExcelColumn('פרטי דרישה', columns=[
            ExcelColumn('חודש'),
            ExcelColumn('סטטוס'),
            ExcelColumn("מס' מכירות"),
            ExcelColumn('סכום דרישה', 'currency', showSum=True)
        ]),
        ExcelColumn('פרטי חשבונית', columns=[
            ExcelColumn("מס' חשבונית"),
            ExcelColumn('סכום'),
            ExcelColumn('תאריך')
        ]),
        ExcelColumn('פרטי שיקים', columns=[
            ExcelColumn('סכום'),
            ExcelColumn('תאריך')
        ]),
        ExcelColumn('הפרשי דרישה', columns=[
            ExcelColumn('דרישה לחשבונית', 'currency', showSum=True),
            ExcelColumn('שיק לחשבונית', 'currency', showSum=True),
            ExcelColumn('זיכוי חשבונית')
        ])
    ]

    rows = []

    title = 'מעקב דרישות - {project_name}'.format(project_name=project.name)

    return ExcelGenerator().generate(title, columns, rows)

class ExcelColumn:
    def __init__(self, title, style=None, showSum=False, width=None, columns=None):
        self.title = title
        self.style = style
        self.showSum = showSum
        self.width = width
        self.columns = columns

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

class ExcelGenerator:

    def __init__(self):

        self.thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin'))

    def _create_workbook(self):
        workbook = Workbook()
        # Get active worksheet/tab
        worksheet = workbook.active
        # set RTL
        worksheet.sheet_view.rightToLeft = True
        
        self.workbook = workbook
        self.worksheet = worksheet

    def _size_columns(self, columns):

        for col_num, col in enumerate(columns):
            if col.width == None:
                continue
            
            col_letter_ascii = ord('A') + col_num
            col_letter = chr(col_letter_ascii)

            self.worksheet.column_dimensions[col_letter].width = col.width

    def _create_title(self, title, columns):
        title_cell = self.worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.style = 'Headline 1'
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        self.worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(columns))

    def _create_header_cell(self, row, column, value):
        cell = self.worksheet.cell(row=row, column=column)
        cell.value = value

        cell.style = 'Accent3'
        cell.border = self.thin_border

    def _create_table_headers(self, columns):
        row_num = 2
        col_num = 1
        has_sub_columns = False

        actual_columns = []

        for column in columns:
            self._create_header_cell(row_num, col_num, column.title)

            if column.columns == None:
                actual_columns.append(column)
                col_num += 1
            else:
                has_sub_columns = True
                for sub_column in column.columns:
                    create_header_cell(row_num + 1, col_num, sub_column.title)
                    actual_columns.append(sub_column)
                    col_num += 1

        # set row_num
        self.row_num = 3 if has_sub_columns else 2
        self._columns = actual_columns

    def _create_table_rows(self, data_rows):
        sum_row = [0 if col.showSum else '' for col in self._columns]
        # override first cell
        sum_row[0] = 'סה"כ'

        for row in data_rows:
            self.row_num += 1

            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = self.worksheet.cell(row=self.row_num, column=col_num)
                cell.value = cell_value if cell_value != None else ''

                col = self._columns[col_num - 1]

                if col.style == 'currency':
                    cell.style = 'Currency'
                    cell.number_format = '#,##0 ₪'
                elif col.style == 'percent':
                    cell.style = 'Percent'
                    # scale cell_value
                    cell.value = cell_value / 100 if cell_value else ''

                cell.border = self.thin_border

                if col.showSum:
                    sum_row[col_num - 1] += cell_value

        self.sum_row = sum_row

    def _create_summary_row(self):
        row_num = self.row_num + 1

        for col_num, cell_value in enumerate(self.sum_row, 1):
            cell = self.worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value or ''

            col = self._columns[col_num - 1]

            if col.style == 'currency':
                cell.style = 'Currency'
                cell.number_format = '#,##0 ₪'

            cell.font = Font(bold=True)
            cell.border = self.thin_border


    def generate(self, title, columns, data_rows):
        self.title = title
        self.columns = columns
        self.data_rows = data_rows

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={title}.xlsx'.format(
            title=title,
        )

        self._create_workbook()

        self._size_columns(columns)

        self._create_title(title, columns)

        self._create_table_headers(columns)

        self._create_table_rows(data_rows)

        self._create_summary_row()

        self.workbook.save(response)

        return response

@permission_required('Management.demand_pay_balance')
def demand_pay_balance_list(request):
    if len(request.GET):
        form = DemandPayBalanceForm(request.GET)
        if form.is_valid():
            # gather form data
            cleaned_data = form.cleaned_data
            demand_pay_balance = cleaned_data['demand_pay_balance']
            project, from_year, from_month, to_year, to_month, all_times = \
                cleaned_data['project'], \
                cleaned_data['from_year'], \
                cleaned_data['from_month'], \
                cleaned_data['to_year'], \
                cleaned_data['to_month'], \
                cleaned_data['all_times']
            
            if project:
                query = Demand.objects.filter(project = project)
            else:
                query = Demand.objects.all()
            if from_year and from_month and to_year and to_month and not all_times:
                query = query.range(from_year, from_month, to_year, to_month)

            all_demands = query \
                .prefetch_related('reminders__statuses', 'invoices__offset','payments') \
                .select_related('project__demand_contact', 'project__payment_contact') \
                .order_by('project_id', 'year', 'month')

            set_demand_sale_fields(all_demands, from_year, from_month, to_year, to_month)
            set_demand_diff_fields(all_demands)
            set_demand_last_status(ds)
            set_demand_invoice_payment_fields(all_demands)
            set_demand_open_reminders(all_demands)

            filters = {
                'un-paid': lambda demand: not demand.payments_amount,
                'mis-paid': lambda demand: demand.diff_invoice and demand.invoices_amount != None,
                'partially-paid': lambda demand: demand.diff_invoice_payment and demand.payments_amount != None,
                'fully-paid': lambda demand: demand.is_fully_paid,
                'all': lambda demand: True,
            }

            filter_func = filters[demand_pay_balance.id]

            # apply a filter to the list of demands
            all_demands = filter(filter_func, all_demands)
            
            # group the demands by project
            project_demands = {}
            
            for project, demand_iter in itertools.groupby(all_demands, lambda demand: demand.project):
                demands = list(demand_iter)
                project_demands[project] = demands

                project.total_amount = sum(map(lambda d: d.total_amount, demands))
                project.total_payments = sum(map(lambda d: d.payments_amount or 0, demands))
                project.total_invoices = sum(map(lambda d: (d.invoices_amount or 0) + (d.invoice_offsets_amount or 0), demands))
                project.total_diff_invoice = sum(map(lambda d: d.diff_invoice, demands))
                project.total_diff_invoice_payment = sum(map(lambda d: d.diff_invoice_payment, demands))
                
            if 'html' in request.GET:
                return render(request, 'Management/demand_pay_balance_list.html', 
                    { 'filterForm': form, 'project_demands': project_demands})
            elif 'pdf' in request.GET:
                # build arguments to the writer
                kwargs = {}
                for key in 'all_times', 'from_month', 'from_year', 'to_month', 'to_year', 'demand_pay_balance':
                    kwargs[key] = cleaned_data[key]
                kwargs['project_demands'] = project_demands
                
                writer = DemandPayBalanceWriter(**kwargs)
                
                return build_and_return_pdf(writer)
    else:
        return render(request, 'Management/demand_pay_balance_list.html', 
                                  { 'filterForm': DemandPayBalanceForm(), 'project_demands': {}},
                                  )

@permission_required('Management.demand_followup')
def demand_followup_list(request):
    ds = Demand.objects.none()
    total_amount, total_invoices, total_payments, total_diff_invoice, total_diff_invoice_payment = 0,0,0,0,0
    from_date, to_date, project = None, None, None
    
    if len(request.GET):
        form = ProjectSeasonForm(request.GET)
        if form.is_valid():
            project = form.cleaned_data['project']

            from_year, from_month = form.cleaned_data['from_year'], form.cleaned_data['from_month']
            to_year, to_month = form.cleaned_data['to_year'], form.cleaned_data['to_month']

            from_date = date(from_year, from_month, 1)
            to_date = date(to_year, to_month, 1)
            
            ds = Demand.objects \
                .range(from_year, from_month, to_year, to_month) \
                .filter(project = project) \
                .select_related('project') \
                .prefetch_related('reminders__statuses', 'invoices__offset','payments')

            set_demand_sale_fields(ds, from_year, from_month, to_year, to_month)
            set_demand_diff_fields(ds)
            set_demand_last_status(ds)
            set_demand_is_fixed(ds)
            set_demand_invoice_payment_fields(ds)
            set_demand_open_reminders(ds)

            for d in ds:
                total_amount += d.total_amount
                total_invoices += d.total_amount_offset
                total_payments += (d.payments_amount or 0)
                total_diff_invoice += d.diff_invoice
                total_diff_invoice_payment += d.diff_invoice_payment
    else:
        form = ProjectSeasonForm()
            
    context = { 
        'demands':ds, 
        'start':from_date, 'end':to_date,
        'project':project, 'filterForm':form,
        'total_amount':total_amount, 
        'total_invoices':total_invoices, 
        'total_payments':total_payments,
        'total_diff_invoice':total_diff_invoice, 
        'total_diff_invoice_payment':total_diff_invoice_payment}

    return render(request, 'Management/demand_followup_list.html', context)

@permission_required('Management.season_employeesalary')
def employeesalary_season_list(request):
    salaries = []
    from_date, to_date, employee_base = None,None,None
    
    total_attrs = ['neto', 'check_amount', 'loan_pay', 'bruto', 'refund']
    totals = dict([('total_' + attr, 0) for attr in total_attrs])
    
    if len(request.GET):
        form = EmployeeSeasonForm(request.GET)
        if form.is_valid():
            employee_base = form.cleaned_data['employee']

            from_year, from_month = form.cleaned_data['from_year'], form.cleaned_data['from_month']
            to_year, to_month = form.cleaned_data['to_year'], form.cleaned_data['to_month']

            from_date = date(from_year, from_month, 1)
            to_date = date(to_year, to_month, 1)

            # set to Employee or NHEmployee
            employee = employee_base.derived

            set_loan_fields([employee])

            if isinstance(employee, Employee):
                salaries = EmployeeSalary.objects.nondeleted() \
                    .range(from_year, from_month, to_year, to_month) \
                    .filter(employee_id = employee_base.id)
                
                # set the same employee instance for all salaries
                for s in salaries:
                    s.employee = employee

                enrich_employee_salaries(
                    salaries, 
                    {employee_base.id: employee},
                    from_year, from_month, to_year, to_month)

            elif isinstance(employee, NHEmployee):
                salaries = NHEmployeeSalary.objects.nondeleted().range(from_year, from_month, to_year, to_month).filter(nhemployee__id = employee_base.id)
                
                # set the same employee instance for all salaries
                for s in salaries:
                    s.nhemployee = employee

                enrich_nh_employee_salaries(
                    salaries, 
                    {employee_base.id: employee},
                    from_year, from_month, to_year, to_month)
            
            if 'list' in request.GET:    
                # aggregate to get total values
                for salary in salaries:
                    for attr in total_attrs:
                        attr_value = getattr(salary, attr)
                        totals['total_' + attr] += attr_value or 0
            elif 'pdf' in request.GET:
                title = u'ריכוז שכר תקופתי לעובד - %s' % employee_base
                writer = EmployeeSalariesWriter(salaries, title, show_month=True, show_employee=False)
                return build_and_return_pdf(writer)
    else:
        form = EmployeeSeasonForm()
    
    context = { 'salaries':salaries, 'start':from_date, 'end':to_date, 'employee':employee_base, 'filterForm':form }
    context.update(totals)
    
    return render(request, 'Management/employeesalary_season_list.html', context, )

@permission_required('Management.season_salaryexpenses')
def employeesalary_season_expenses(request):
    salaries = []
    total_neto, total_check_amount, total_loan_pay, total_bruto, total_bruto_employer, total_refund = 0,0,0,0,0,0
    total_sale_count = 0
    from_date, to_date, employee_base = None,None,None
    
    if len(request.GET):
        form = EmployeeSeasonForm(request.GET)
        if form.is_valid():
            employee_base = form.cleaned_data['employee']
            from_date = date(form.cleaned_data['from_year'], form.cleaned_data['from_month'], 1)
            to_date = date(form.cleaned_data['to_year'], form.cleaned_data['to_month'], 1)

            employee = employee_base.derived

            if isinstance(employee, Employee):
                salaries = EmployeeSalary.objects.nondeleted() \
                    .range(from_date.year, from_date.month, to_date.year, to_date.month) \
                    .select_related('employee__employment_terms__hire_type') \
                    .filter(employee_id = employee_base.id)
                
                enrich_employee_salaries(
                    salaries, 
                    {employee_base.id: employee},
                    from_date.year, from_date.month, to_date.year, to_date.month)
                
                template = 'Management/employeesalary_season_expenses.html'
            elif isinstance(employee, NHEmployee):
                salaries = NHEmployeeSalary.objects.nondeleted().range(from_date.year, from_date.month, to_date.year, to_date.month).filter(nhemployee__id = employee_base.id)

                enrich_nh_employee_salaries(
                    salaries, 
                    {employee_base.id: employee},
                    from_date.year, from_date.month, to_date.year, to_date.month)

                template = 'Management/nhemployeesalary_season_expenses.html'

            for salary in salaries:
                total_neto += salary.neto or 0
                total_check_amount += salary.check_amount or 0
                total_loan_pay += salary.loan_pay or 0
                total_bruto += salary.bruto or 0
                total_bruto_employer += salary.bruto_employer_expense or 0
    else:
        template = 'Management/employeesalary_season_expenses.html'
        form = EmployeeSeasonForm()
        
    return render(request, template, 
                              { 'salaries':salaries, 'start':from_date, 'end':to_date,
                                'employee': employee_base, 'filterForm':form,
                                'total_neto':total_neto,'total_check_amount':total_check_amount,
                                'total_loan_pay':total_loan_pay,'total_bruto':total_bruto,'total_bruto_employer':total_bruto_employer},
                              )

@permission_required('Management.season_total_salaryexpenses')
def employeesalary_season_total_expenses(request):
    employees = []
    to_date, from_date = None, None
    
    if len(request.GET):
        form = DivisionTypeSeasonForm(request.GET)
        if form.is_valid():
            # extract form values
            division_type = form.cleaned_data['division_type']

            from_year, from_month = form.cleaned_data['from_year'], form.cleaned_data['from_month']
            to_year, to_month = form.cleaned_data['to_year'], form.cleaned_data['to_month']

            from_date = date(from_year, from_month, 1)
            to_date = date(to_year, to_month, 1)

            if division_type.id == DivisionType.Marketing:
                employees = list(Employee.objects \
                    .select_related('employment_terms') \
                    .exclude(work_end__isnull = False, work_end__lt = from_date))

                employee_by_id = {e.id:e for e in employees}
                salaries = EmployeeSalary.objects \
                    .select_related('employee__employment_terms') \
                    .range(from_year, from_month, to_year, to_month) \
                    .order_by('employee_id')
            else:
                # NiceHouse division

                # map DivisionType to NHBranch
                division_to_branch = {
                    DivisionType.NHShoham: NHBranch.Shoham,
                    DivisionType.NHModiin: NHBranch.Modiin,
                    DivisionType.NHNesZiona: NHBranch.NesZiona,
                }

                branch_id = division_to_branch[division_type.id]
                
                query = NHBranchEmployee.objects \
                    .select_related('nhemployee__employment_terms') \
                    .filter(nhbranch__id = branch_id) \
                    .exclude(end_date__isnull=False, end_date__lt = from_date)

                employees = [x.nhemployee for x in query]
                employee_by_id = {e.id:e for e in employees}
                salaries = NHEmployeeSalary.objects \
                    .range(from_year, from_month, to_year, to_month) \
                    .order_by('nhemployee_id')

            set_salary_base_fields(
                salaries,
                employee_by_id,
                from_year, from_month, to_year, to_month)

            attrs = ['neto', 'loan_pay', 'check_amount', 'income_tax', 'national_insurance', 'health', 'pension_insurance', 
                     'vacation', 'convalescence_pay', 'bruto', 'employer_national_insurance', 'employer_benefit',
                     'compensation_allocation', 'bruto_with_employer']
            
            for employee_id, employee_salaries in itertools.groupby(salaries, lambda salary: salary.get_employee().id):
                # get the employee object
                employee = employee_by_id[employee_id]
                
                # evaluate and store the iterator
                employee_salaries_list = list(employee_salaries)

                for attr in attrs:
                    # extract 'attr' from each salary
                    attr_list = [getattr(salary, attr, 0) or 0 for salary in employee_salaries_list]
                    # sum 'attr' over employee_salaries
                    attr_sum = sum(attr_list)
                    # set 'total_' attribute on employee object
                    setattr(employee, 'total_' + attr, attr_sum)

    else:
        form = DivisionTypeSeasonForm()
            
    context = { 
        'employees':employees, 
        'start':from_date, 
        'end':to_date, 
        'filterForm':form 
    }

    return render(request, 'Management/employeesalary_season_total_expenses.html', context)